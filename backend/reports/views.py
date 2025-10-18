"""
MonthlyReport ViewSet - RESTful API for Monthly Report management
"""

from django.http import HttpResponse
from django.db import transaction
from django.db.models import Count, Sum
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

from chapters.models import Chapter
from chapters.permissions import IsChapterOrAdmin, IsAdmin
from members.models import Member
from reports.models import MonthlyReport, MemberMonthlyStats
from analytics.models import TYFCB
from bni.services.aggregation_service import AggregationService

# Import new modular formatters
from bni.services.excel_formatters import (
    write_referral_matrix,
    write_oto_matrix,
    write_combination_matrix,
    write_tyfcb_report,
)


class MonthlyReportViewSet(viewsets.ModelViewSet):
    """
    ViewSet for MonthlyReport operations.

    All endpoints require authentication. Chapters can only access their own reports, admins can access all.

    Provides:
    - list: Get all monthly reports for a chapter (IsChapterOrAdmin)
    - destroy: Delete a monthly report (IsAdmin)
    - member_detail: Get detailed member information (IsChapterOrAdmin)
    - tyfcb_data: Get TYFCB data for a specific report (IsChapterOrAdmin)
    - download_matrices: Download matrices as Excel (IsChapterOrAdmin)
    - download_palms: Download original PALMS sheets (IsChapterOrAdmin)
    - aggregate_reports: Aggregate multiple reports (IsChapterOrAdmin)
    - download_aggregated: Download aggregated analysis (IsChapterOrAdmin)
    """

    queryset = MonthlyReport.objects.all()
    permission_classes = [IsChapterOrAdmin]  # Chapters see their own, admins see all

    def get_permissions(self):
        """Override permissions based on action."""
        if self.action == "destroy":
            # Only admins can delete reports
            return [IsAdmin()]
        return [IsChapterOrAdmin()]

    def list(self, request, chapter_id=None):
        """
        Get all monthly reports for a specific chapter.

        Returns list of monthly reports with metadata including:
        - Report IDs and dates
        - File information
        - Matrix availability flags
        """
        try:
            chapter = Chapter.objects.get(id=chapter_id)
            monthly_reports = MonthlyReport.objects.filter(chapter=chapter).order_by(
                "-month_year"
            )

            result = []
            for report in monthly_reports:
                try:
                    # Safely access file fields (they are strings, not actual files)
                    slip_file = (
                        str(report.slip_audit_file) if report.slip_audit_file else None
                    )
                    member_file = (
                        str(report.member_names_file)
                        if report.member_names_file
                        else None
                    )

                    result.append(
                        {
                            "id": report.id,
                            "month_year": report.month_year,
                            "uploaded_at": report.uploaded_at.isoformat()
                            if report.uploaded_at
                            else None,
                            "processed_at": report.processed_at.isoformat()
                            if report.processed_at
                            else None,
                            "slip_audit_file": slip_file,
                            "member_names_file": member_file,
                            "has_referral_matrix": bool(report.referral_matrix_data),
                            "has_oto_matrix": bool(report.oto_matrix_data),
                            "has_combination_matrix": bool(
                                report.combination_matrix_data
                            ),
                            "week_of_date": report.week_of_date.isoformat()
                            if report.week_of_date
                            else None,
                            "audit_period_start": report.audit_period_start.isoformat()
                            if report.audit_period_start
                            else None,
                            "audit_period_end": report.audit_period_end.isoformat()
                            if report.audit_period_end
                            else None,
                            "require_palms_sheets": report.require_palms_sheets,
                            "uploaded_file_names": report.uploaded_file_names,
                        }
                    )
                except Exception as e:
                    # Log the error but continue processing other reports
                    logger.error(f"Error processing report {report.id}: {str(e)}")
                    continue

            return Response(result)

        except Chapter.DoesNotExist:
            return Response(
                {"error": "Chapter not found"}, status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": f"Monthly reports retrieval failed: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @transaction.atomic
    def destroy(self, request, pk=None, chapter_id=None):
        """
        Delete a monthly report atomically.

        Uses @transaction.atomic decorator to ensure all cascade deletions
        (MemberMonthlyStats, referrals, etc.) complete successfully or rollback entirely.
        This prevents orphaned records if deletion fails partway through.

        Authentication is handled by IsAdmin permission class.
        """
        try:
            chapter = Chapter.objects.get(id=chapter_id)
            monthly_report = MonthlyReport.objects.get(id=pk, chapter=chapter)

            # Delete the report (files are just filenames stored as strings, no actual files to delete)
            # Transaction ensures all cascade deletions succeed or rollback:
            # - MemberMonthlyStats (via foreign key cascade)
            # - Associated analytics data
            monthly_report.delete()

            return Response({"message": "Monthly report deleted successfully"})

        except (Chapter.DoesNotExist, MonthlyReport.DoesNotExist):
            return Response(
                {"error": "Chapter or monthly report not found"},
                status=status.HTTP_404_NOT_FOUND,
            )
        except Exception as e:
            logger.exception(f"Failed to delete report {pk}")
            return Response(
                {"error": f"Failed to delete report: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=True, methods=["get"], url_path="members/(?P<member_id>[^/.]+)")
    def member_detail(self, request, pk=None, chapter_id=None, member_id=None):
        """
        Get detailed member information including missing interaction lists.

        Returns member stats and lists of:
        - Missing one-to-ones
        - Missing referrals (given and received)
        - Priority connections (members appearing in multiple missing lists)
        """
        try:
            chapter = Chapter.objects.get(id=chapter_id)
            monthly_report = MonthlyReport.objects.get(id=pk, chapter=chapter)
            member = Member.objects.get(id=member_id, chapter=chapter)

            try:
                member_stats = MemberMonthlyStats.objects.get(
                    member=member, monthly_report=monthly_report
                )
            except MemberMonthlyStats.DoesNotExist:
                # If no stats exist, return basic member info with empty lists
                member_stats = None

            # Get all chapter members for name resolution
            chapter_members = Member.objects.filter(chapter=chapter, is_active=True)
            member_lookup = {m.id: m.full_name for m in chapter_members}

            result = {
                "member": {
                    "id": member.id,
                    "full_name": member.full_name,
                    "first_name": member.first_name,
                    "last_name": member.last_name,
                    "business_name": member.business_name,
                    "classification": member.classification,
                    "email": member.email,
                    "phone": member.phone,
                },
                "stats": {
                    "referrals_given": member_stats.referrals_given
                    if member_stats
                    else 0,
                    "referrals_received": member_stats.referrals_received
                    if member_stats
                    else 0,
                    "one_to_ones_completed": member_stats.one_to_ones_completed
                    if member_stats
                    else 0,
                    "tyfcb_inside_amount": float(member_stats.tyfcb_inside_amount)
                    if member_stats
                    else 0.0,
                    "tyfcb_outside_amount": float(member_stats.tyfcb_outside_amount)
                    if member_stats
                    else 0.0,
                },
                "missing_interactions": {
                    "missing_otos": [
                        {
                            "id": member_id,
                            "name": member_lookup.get(member_id, "Unknown"),
                        }
                        for member_id in (
                            member_stats.missing_otos if member_stats else []
                        )
                        if member_id in member_lookup
                    ],
                    "missing_referrals_given_to": [
                        {
                            "id": member_id,
                            "name": member_lookup.get(member_id, "Unknown"),
                        }
                        for member_id in (
                            member_stats.missing_referrals_given_to
                            if member_stats
                            else []
                        )
                        if member_id in member_lookup
                    ],
                    "missing_referrals_received_from": [
                        {
                            "id": member_id,
                            "name": member_lookup.get(member_id, "Unknown"),
                        }
                        for member_id in (
                            member_stats.missing_referrals_received_from
                            if member_stats
                            else []
                        )
                        if member_id in member_lookup
                    ],
                    "priority_connections": [
                        {
                            "id": member_id,
                            "name": member_lookup.get(member_id, "Unknown"),
                        }
                        for member_id in (
                            member_stats.priority_connections if member_stats else []
                        )
                        if member_id in member_lookup
                    ],
                },
                "monthly_report": {
                    "id": monthly_report.id,
                    "month_year": monthly_report.month_year,
                    "processed_at": monthly_report.processed_at,
                },
            }

            return Response(result)

        except (Chapter.DoesNotExist, MonthlyReport.DoesNotExist, Member.DoesNotExist):
            return Response(
                {"error": "Chapter, monthly report, or member not found"},
                status=status.HTTP_404_NOT_FOUND,
            )
        except Exception as e:
            return Response(
                {"error": f"Member detail retrieval failed: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=True, methods=["get"], url_path="tyfcb")
    def tyfcb_data(self, request, pk=None, chapter_id=None):
        """
        Get TYFCB data for a specific monthly report.

        Returns inside and outside TYFCB data with totals and per-member breakdowns.
        """
        try:
            chapter = Chapter.objects.get(id=chapter_id)
            monthly_report = MonthlyReport.objects.get(id=pk, chapter=chapter)

            tyfcb_data = {
                "inside": monthly_report.tyfcb_inside_data
                or {"total_amount": 0, "count": 0, "by_member": {}},
                "outside": monthly_report.tyfcb_outside_data
                or {"total_amount": 0, "count": 0, "by_member": {}},
                "month_year": monthly_report.month_year,
                "processed_at": monthly_report.processed_at,
            }

            return Response(tyfcb_data)

        except Chapter.DoesNotExist:
            return Response(
                {"error": "Chapter not found"}, status=status.HTTP_404_NOT_FOUND
            )
        except MonthlyReport.DoesNotExist:
            return Response(
                {"error": "Monthly report not found"}, status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": f"Failed to get TYFCB data: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=True, methods=["get"], url_path="download-matrices")
    def download_matrices(self, request, pk=None, chapter_id=None):
        """
        Generate and download Excel file with all matrices.

        Creates a workbook with separate sheets for:
        - Referral Matrix
        - One-to-One Matrix
        - Combination Matrix
        - TYFCB Report

        Uses modular formatters for consistent styling across single-month and multi-month reports.
        """
        try:
            chapter = Chapter.objects.get(id=chapter_id)
            monthly_report = MonthlyReport.objects.get(id=pk, chapter=chapter)

            # Import pandas for DataFrame conversion
            import pandas as pd

            # Create a new workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Prepare data for formatters - convert single month to format expected by formatters
            # Period string for single month
            from datetime import datetime
            try:
                month_date = datetime.strptime(monthly_report.month_year, "%Y-%m")
                period_str = month_date.strftime("%m/%Y")
            except:
                period_str = monthly_report.month_year

            # Create a temporary list with this single report for the formatters
            reports_list = [monthly_report]

            # Calculate chapter statistics (needed for performance highlighting)
            # Build aggregated data from this single report
            if monthly_report.referral_matrix_data and "members" in monthly_report.referral_matrix_data:
                members = monthly_report.referral_matrix_data["members"]
                ref_matrix_data = monthly_report.referral_matrix_data["matrix"]

                # Convert to DataFrame
                ref_df = pd.DataFrame(ref_matrix_data, index=members, columns=members)

                # Calculate statistics
                ref_totals = ref_df.sum(axis=1).to_dict()
                avg_referrals = ref_df.sum(axis=1).mean() if len(members) > 0 else 0
            else:
                ref_df = pd.DataFrame()
                ref_totals = {}
                avg_referrals = 0

            if monthly_report.oto_matrix_data and "members" in monthly_report.oto_matrix_data:
                members_oto = monthly_report.oto_matrix_data["members"]
                oto_matrix_data = monthly_report.oto_matrix_data["matrix"]
                oto_df = pd.DataFrame(oto_matrix_data, index=members_oto, columns=members_oto)
                oto_totals = oto_df.sum(axis=1).to_dict()
                avg_oto = oto_df.sum(axis=1).mean() if len(members_oto) > 0 else 0
            else:
                oto_df = pd.DataFrame()
                oto_totals = {}
                avg_oto = 0

            if monthly_report.combination_matrix_data and "members" in monthly_report.combination_matrix_data:
                members_combo = monthly_report.combination_matrix_data["members"]
                combo_matrix_data = monthly_report.combination_matrix_data["matrix"]
                combo_df = pd.DataFrame(combo_matrix_data, index=members_combo, columns=members_combo)
            else:
                combo_df = pd.DataFrame()

            # Calculate TYFCB statistics
            tyfcb_totals = {}
            tyfcb_inside = monthly_report.tyfcb_inside_data or {}
            tyfcb_outside = monthly_report.tyfcb_outside_data or {}

            if tyfcb_inside or tyfcb_outside:
                all_members = set()
                if tyfcb_inside and "by_member" in tyfcb_inside:
                    all_members.update(tyfcb_inside["by_member"].keys())
                if tyfcb_outside and "by_member" in tyfcb_outside:
                    all_members.update(tyfcb_outside["by_member"].keys())

                for member in all_members:
                    inside_amt = float(tyfcb_inside.get("by_member", {}).get(member, 0))
                    outside_amt = float(tyfcb_outside.get("by_member", {}).get(member, 0))
                    tyfcb_totals[member] = inside_amt + outside_amt

            avg_tyfcb = sum(tyfcb_totals.values()) / len(tyfcb_totals) if tyfcb_totals else 0

            # Build stats dict for formatters
            stats = {
                "chapter_size": len(members) if monthly_report.referral_matrix_data else 0,
                "avg_referrals": avg_referrals,
                "avg_oto": avg_oto,
                "avg_tyfcb": avg_tyfcb,
                "ref_totals": ref_totals,
                "oto_totals": oto_totals,
                "tyfcb_totals": tyfcb_totals,
                "total_months": 1,
            }

            # Create sheets using the new formatters
            if monthly_report.referral_matrix_data and not ref_df.empty:
                ws_ref = wb.create_sheet("Referral Matrix")
                write_referral_matrix(ws_ref, ref_df, period_str, stats, reports_list)

            if monthly_report.oto_matrix_data and not oto_df.empty:
                ws_oto = wb.create_sheet("One-to-One Matrix")
                write_oto_matrix(ws_oto, oto_df, period_str, stats, reports_list)

            if monthly_report.combination_matrix_data and not combo_df.empty:
                ws_combo = wb.create_sheet("Combination Matrix")
                write_combination_matrix(ws_combo, combo_df, period_str, stats, reports_list)

            # Create TYFCB sheet using the new formatter
            if tyfcb_inside or tyfcb_outside:
                ws_tyfcb = wb.create_sheet("TYFCB Report")
                # Extract by_member dicts for the formatter
                inside_by_member = tyfcb_inside.get("by_member", {}) if tyfcb_inside else {}
                outside_by_member = tyfcb_outside.get("by_member", {}) if tyfcb_outside else {}
                write_tyfcb_report(ws_tyfcb, inside_by_member, outside_by_member, period_str, stats, reports_list)

            # If no matrices were created, add an info sheet
            if len(wb.sheetnames) == 0:
                ws = wb.create_sheet("Info")
                ws["A1"] = "No matrix data available for this report"
                ws["A1"].font = Font(bold=True)

            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            # Create HTTP response
            filename = f"{chapter.name.replace(' ', '_')}_Matrices_{monthly_report.month_year}.xlsx"
            response = HttpResponse(
                output.getvalue(),  # Use getvalue() instead of read() to avoid buffer position issues
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response

        except Chapter.DoesNotExist:
            return Response(
                {"error": "Chapter not found"}, status=status.HTTP_404_NOT_FOUND
            )
        except MonthlyReport.DoesNotExist:
            return Response(
                {"error": "Monthly report not found"}, status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": f"Failed to generate Excel file: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=True, methods=["get"], url_path="download-palms")
    def download_palms(self, request, pk=None, chapter_id=None):
        """
        Download original PALMS sheets for a report.

        Returns:
        - ZIP file containing all uploaded PALMS sheets if require_palms_sheets is True
        - Error message if PALMS sheets are not available
        """
        import zipfile
        import os
        from django.conf import settings

        try:
            chapter = Chapter.objects.get(id=chapter_id)
            monthly_report = MonthlyReport.objects.get(id=pk, chapter=chapter)

            # Check if PALMS sheets are marked as downloadable
            if not monthly_report.require_palms_sheets:
                return Response(
                    {
                        "error": "PALMS sheets were not marked as downloadable for this report"
                    },
                    status=status.HTTP_404_NOT_FOUND,
                )

            # Check if we have file metadata
            if not monthly_report.uploaded_file_names:
                return Response(
                    {"error": "No file metadata available for this report"},
                    status=status.HTTP_404_NOT_FOUND,
                )

            # Create ZIP file in memory
            zip_buffer = BytesIO()
            files_found = 0

            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                # Add each PALMS file to the ZIP
                for file_info in monthly_report.uploaded_file_names:
                    if file_info.get("file_type") == "slip_audit":
                        # Use saved_filename if available (new format), otherwise fall back to original_filename
                        saved_filename = file_info.get("saved_filename")
                        original_filename = file_info.get(
                            "original_filename", "slip_audit.xls"
                        )
                        file_path_rel = file_info.get(
                            "file_path"
                        )  # e.g., "uploads/filename.xls"

                        media_root = getattr(settings, "MEDIA_ROOT", None)
                        if not media_root:
                            zip_file.writestr(
                                "README.txt",
                                "File storage not configured. Original files may not be available.\n",
                            )
                            continue

                        # Try using the stored file_path first (most reliable)
                        if file_path_rel:
                            file_path = os.path.join(media_root, file_path_rel)
                            if os.path.exists(file_path):
                                with open(file_path, "rb") as f:
                                    zip_file.writestr(original_filename, f.read())
                                files_found += 1
                                continue

                        # Fallback: try using saved_filename
                        if saved_filename:
                            file_path = os.path.join(
                                media_root, "uploads", saved_filename
                            )
                            if os.path.exists(file_path):
                                with open(file_path, "rb") as f:
                                    zip_file.writestr(original_filename, f.read())
                                files_found += 1
                                continue

                        # Last resort: try original filename (backward compatibility for old uploads)
                        file_path = os.path.join(
                            media_root, "uploads", original_filename
                        )
                        if os.path.exists(file_path):
                            with open(file_path, "rb") as f:
                                zip_file.writestr(original_filename, f.read())
                            files_found += 1

                # If no files were found, add an error message
                if files_found == 0:
                    zip_file.writestr(
                        "ERROR.txt",
                        "No PALMS files were found on the server.\n\n"
                        "This could mean:\n"
                        "1. The files were uploaded before file storage was implemented\n"
                        "2. The files were deleted from the server\n"
                        "3. The upload did not complete successfully\n\n"
                        "Please re-upload the report with PALMS sheets enabled.",
                    )

            zip_buffer.seek(0)

            # Generate filename with date
            date_str = (
                monthly_report.week_of_date.strftime("%Y-%m-%d")
                if monthly_report.week_of_date
                else monthly_report.month_year
            )
            filename = f"PALMS_Sheets_{chapter.name.replace(' ', '_')}_{date_str}.zip"

            response = HttpResponse(
                zip_buffer.getvalue(),
                content_type="application/zip",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response

        except Chapter.DoesNotExist:
            return Response(
                {"error": "Chapter not found"}, status=status.HTTP_404_NOT_FOUND
            )
        except MonthlyReport.DoesNotExist:
            return Response(
                {"error": "Monthly report not found"}, status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.exception(f"Failed to download PALMS files for report {pk}")
            return Response(
                {"error": f"Failed to download PALMS files: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["post"], url_path="aggregate")
    def aggregate_reports(self, request, chapter_id=None):
        """
        Aggregate multiple monthly reports into combined analysis.

        Request body:
            {
                "report_ids": [1, 2, 3, ...]
            }

        Returns:
            - Aggregated matrices (referral, OTO, combination)
            - TYFCB data
            - Member completeness information
            - Month range and metadata
        """
        try:
            chapter = Chapter.objects.get(id=chapter_id)
            report_ids = request.data.get("report_ids", [])

            if not report_ids:
                return Response(
                    {"error": "No report IDs provided"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Fetch reports
            reports = MonthlyReport.objects.filter(
                id__in=report_ids, chapter=chapter
            ).order_by("month_year")

            if not reports.exists():
                return Response(
                    {"error": "No valid reports found"},
                    status=status.HTTP_404_NOT_FOUND,
                )

            # Create aggregation service and process
            aggregation_service = AggregationService(list(reports))
            aggregated_data = aggregation_service.aggregate_matrices()
            member_differences = aggregation_service.get_member_differences()

            # Add member differences to response
            aggregated_data["member_differences"] = member_differences

            return Response(aggregated_data)

        except Chapter.DoesNotExist:
            return Response(
                {"error": "Chapter not found"}, status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.exception(f"Aggregation failed for chapter {chapter_id}")
            return Response(
                {"error": f"Aggregation failed: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["post"], url_path="aggregate/download")
    def download_aggregated(self, request, chapter_id=None):
        """
        Download comprehensive Excel file with aggregated analysis.

        Request body:
            {
                "report_ids": [1, 2, 3, ...]
            }

        Returns:
            Single Excel file containing:
            - Summary sheet
            - Aggregated matrices (Referral, OTO, Combination)
            - TYFCB data (Inside & Outside)
            - Member differences (Inactive members)
        """
        try:
            chapter = Chapter.objects.get(id=chapter_id)
            report_ids = request.data.get("report_ids", [])

            if not report_ids:
                return Response(
                    {"error": "No report IDs provided"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Fetch reports
            reports = MonthlyReport.objects.filter(
                id__in=report_ids, chapter=chapter
            ).order_by("month_year")

            if not reports.exists():
                return Response(
                    {"error": "No valid reports found"},
                    status=status.HTTP_404_NOT_FOUND,
                )

            # Create aggregation service and generate comprehensive Excel file
            aggregation_service = AggregationService(list(reports))
            excel_buffer = aggregation_service.generate_download_package()

            # Create HTTP response with Excel file
            month_range = aggregation_service._get_month_range()
            filename = (
                f"{chapter.name.replace(' ', '_')}_Aggregated_Report_{month_range}.xlsx"
            )

            response = HttpResponse(
                excel_buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            return response

        except Chapter.DoesNotExist:
            return Response(
                {"error": "Chapter not found"}, status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.exception(f"Download generation failed for chapter {chapter_id}")
            return Response(
                {"error": f"Download generation failed: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
"""
File Upload ViewSet - RESTful API for Excel file uploads

Authentication:
- Excel upload: IsChapterOrAdmin (chapters upload their own data)
- Bulk upload: IsAdmin (admin-only operation)
- Reset all: IsAdmin (admin-only operation)
"""

import logging
import re
from datetime import datetime
from rest_framework import viewsets, status, serializers
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser

from chapters.models import Chapter
from chapters.permissions import IsChapterOrAdmin, IsAdmin
from bni.services.excel_processor import ExcelProcessorService
from bni.services.bulk_upload_service import BulkUploadService

logger = logging.getLogger(__name__)


class FileUploadSerializer(serializers.Serializer):
    """Serializer for file upload validation."""

    slip_audit_files = serializers.ListField(
        child=serializers.FileField(),
        allow_empty=False,
        help_text="One or more slip audit files to compile into a single monthly report",
    )
    member_names_file = serializers.FileField(required=False)
    chapter_id = serializers.IntegerField()
    month_year = serializers.CharField(
        max_length=7,
        required=False,
        allow_blank=True,
        help_text="e.g., '2024-06' (optional, will be extracted from filename if not provided)",
    )
    upload_option = serializers.ChoiceField(
        choices=["slip_only", "slip_and_members"], default="slip_only"
    )
    week_of_date = serializers.DateField(
        required=False,
        allow_null=True,
        help_text="The week this audit represents (e.g., '2025-01-28')",
    )
    require_palms_sheets = serializers.BooleanField(
        required=False,
        default=False,
        help_text="Whether PALMS data sheets are required",
    )


class FileUploadViewSet(viewsets.ViewSet):
    """
    ViewSet for file upload operations.

    Provides endpoints for:
    - Excel file upload: Process slip audit and member names files (IsChapterOrAdmin)
    - Bulk upload: Process Regional PALMS Summary reports (IsAdmin)
    - Reset all: Delete all data (IsAdmin)

    Permissions:
    - Chapters can upload their own data
    - Admins have full access to all operations
    """

    parser_classes = (MultiPartParser, FormParser)
    permission_classes = [IsChapterOrAdmin]

    def get_permissions(self):
        """Override permissions based on action."""
        if self.action in ['bulk_upload', 'reset_all_data']:
            return [IsAdmin()]  # Admin-only operations
        return [IsChapterOrAdmin()]  # Regular uploads

    def _extract_date_from_filename(self, filename):
        """
        Extract date from slip audit report filename.

        Supports multiple formats:
        - YYYY-MM-DD: slips-audit-report_2025-01-28.xls
        - MM-DD-YYYY: Slips_Audit_Report_08-25-2025_2-26_PM.xls

        Returns month_year in format 'YYYY-MM' or None if not found
        """
        # Try YYYY-MM-DD format first (e.g., 2025-01-28)
        pattern_ymd = r"(\d{4})-(\d{2})-(\d{2})"
        match = re.search(pattern_ymd, filename)
        if match:
            year, month, day = match.groups()
            return f"{year}-{month}"

        # Try MM-DD-YYYY format (e.g., 08-25-2025)
        pattern_mdy = r"(\d{2})-(\d{2})-(\d{4})"
        match = re.search(pattern_mdy, filename)
        if match:
            month, day, year = match.groups()
            return f"{year}-{month}"

        return None

    def _extract_full_date_from_filename(self, filename):
        """
        Extract full date from slip audit report filename.

        Supports multiple formats:
        - YYYY-MM-DD: slips-audit-report_2025-01-28.xls
        - MM-DD-YYYY: Slips_Audit_Report_08-25-2025_2-26_PM.xls

        Returns date object or None if not found
        """
        from datetime import date

        # Try YYYY-MM-DD format first (e.g., 2025-01-28)
        pattern_ymd = r"(\d{4})-(\d{2})-(\d{2})"
        match = re.search(pattern_ymd, filename)
        if match:
            year, month, day = match.groups()
            try:
                return date(int(year), int(month), int(day))
            except ValueError:
                pass

        # Try MM-DD-YYYY format (e.g., 08-25-2025)
        pattern_mdy = r"(\d{2})-(\d{2})-(\d{4})"
        match = re.search(pattern_mdy, filename)
        if match:
            month, day, year = match.groups()
            try:
                return date(int(year), int(month), int(day))
            except ValueError:
                pass

        return None

    @action(detail=False, methods=["post"], url_path="excel")
    def upload_excel(self, request):
        """
        Handle Excel file upload and processing.

        Accepts:
        - slip_audit_file: Required Excel file (.xls/.xlsx)
        - member_names_file: Optional Excel file for member names
        - chapter_id: Chapter to associate with
        - month_year: Optional report month in format 'YYYY-MM' (defaults to current month)
        - upload_option: 'slip_only' or 'slip_and_members'

        Returns processing result with created records and any errors.
        """
        # Log incoming request for debugging
        logger.info(
            f"Excel upload request - Files: {list(request.FILES.keys())}, Data: {list(request.data.keys())}"
        )

        # Handle both new (slip_audit_files) and old (slip_audit_file) format
        slip_files = request.FILES.getlist("slip_audit_files")
        if not slip_files:
            # Try old format for backward compatibility
            single_file = request.FILES.get("slip_audit_file")
            if single_file:
                slip_files = [single_file]

        if not slip_files:
            return Response(
                {"error": "No slip audit files provided"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        logger.info(
            f"Found {len(slip_files)} slip audit file(s): {[f.name for f in slip_files]}"
        )

        # Get other data from request directly (skip serializer for file lists)
        try:
            slip_audit_files = slip_files
            member_names_file = request.FILES.get("member_names_file")

            # Validate required fields
            if not request.data.get("chapter_id"):
                return Response(
                    {"error": "chapter_id is required"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            chapter_id = int(request.data.get("chapter_id"))
            month_year = request.data.get("month_year", "").strip() or None
            upload_option = request.data.get("upload_option", "slip_only")
            week_of_date = request.data.get("week_of_date", "").strip() or None
            require_palms_sheets = request.data.get(
                "require_palms_sheets", "false"
            ).lower() in ["true", "1", "yes"]
        except ValueError as e:
            return Response(
                {"error": f"Invalid data format: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            logger.info(f"Processing {len(slip_audit_files)} slip audit file(s)")

            # If month_year not provided, try to extract from first filename (optional, non-blocking)
            if not month_year:
                extracted_date = self._extract_date_from_filename(
                    slip_audit_files[0].name
                )
                if extracted_date:
                    month_year = extracted_date
                    logger.info(f"Extracted date from filename: {month_year}")
                else:
                    # Use current month as default if extraction fails
                    from datetime import datetime

                    month_year = datetime.now().strftime("%Y-%m")
                    logger.info(f"Using current month as default: {month_year}")

            # If week_of_date not provided, try to extract full date from first filename
            if not week_of_date:
                extracted_full_date = self._extract_full_date_from_filename(
                    slip_audit_files[0].name
                )
                if extracted_full_date:
                    week_of_date = extracted_full_date
                    logger.info(f"Extracted week date from filename: {week_of_date}")

            # Convert week_of_date string to date object if needed
            if week_of_date and isinstance(week_of_date, str):
                from datetime import datetime

                try:
                    week_of_date = datetime.strptime(week_of_date, "%Y-%m-%d").date()
                except ValueError:
                    logger.warning(
                        f"Invalid week_of_date format: {week_of_date}, ignoring"
                    )
                    week_of_date = None

            # Validate chapter access
            try:
                chapter = Chapter.objects.get(id=chapter_id)
                # Add permission check here if needed
            except Chapter.DoesNotExist:
                return Response(
                    {"error": "Chapter not found"}, status=status.HTTP_404_NOT_FOUND
                )

            # Validate file types
            for slip_file in slip_audit_files:
                if not slip_file.name.lower().endswith((".xls", ".xlsx")):
                    return Response(
                        {
                            "error": f"Only .xls and .xlsx files are supported. Invalid file: {slip_file.name}"
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            if member_names_file and not member_names_file.name.lower().endswith(
                (".xls", ".xlsx")
            ):
                return Response(
                    {
                        "error": "Only .xls and .xlsx files are supported for member names file"
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Process using the new monthly report method
            logger.info(
                f"Starting Excel processing for chapter {chapter.name}, month {month_year}"
            )
            logger.info(
                f"Processing {len(slip_audit_files)} slip audit files: {[f.name for f in slip_audit_files]}"
            )
            processor = ExcelProcessorService(chapter)

            try:
                result = processor.process_monthly_reports_batch(
                    slip_audit_files=slip_audit_files,
                    member_names_file=member_names_file,
                    month_year=month_year,
                    week_of_date=week_of_date,
                    require_palms_sheets=require_palms_sheets,
                )

                logger.info(f"Processing complete - Success: {result.get('success')}")

                # Return appropriate status code based on result
                if result.get("success"):
                    return Response(result, status=status.HTTP_200_OK)
                else:
                    logger.error(
                        f"Processing failed: {result.get('error', 'Unknown error')}"
                    )
                    return Response(result, status=status.HTTP_400_BAD_REQUEST)

            except Exception as proc_error:
                logger.exception(f"Excel processing error: {str(proc_error)}")
                return Response(
                    {
                        "error": f"Excel processing failed: {str(proc_error)}",
                        "success": False,
                    },
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

        except Exception as e:
            logger.exception(f"Upload endpoint error: {str(e)}")
            return Response(
                {"error": f"Upload failed: {str(e)}", "success": False},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["post"], url_path="bulk")
    def bulk_upload(self, request):
        """
        Handle Regional PALMS Summary bulk upload.

        Accepts:
        - file: Excel file containing regional summary data

        Processes the file and creates/updates multiple chapters and members.
        Returns summary of operations performed.
        """
        if "file" not in request.FILES:
            return Response(
                {"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST
            )

        file = request.FILES["file"]

        # Validate file type
        if not file.name.endswith((".xls", ".xlsx")):
            return Response(
                {"error": "Invalid file type. Please upload .xls or .xlsx file"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            # Process the file
            service = BulkUploadService()
            result = service.process_region_summary(file)

            if result["success"]:
                return Response(
                    {
                        "success": True,
                        "message": "Bulk upload completed successfully",
                        "details": {
                            "chapters_created": result["chapters_created"],
                            "chapters_updated": result["chapters_updated"],
                            "members_created": result["members_created"],
                            "members_updated": result["members_updated"],
                            "total_processed": result["total_processed"],
                            "warnings": result["warnings"],
                        },
                    },
                    status=status.HTTP_200_OK,
                )
            else:
                return Response(
                    {
                        "success": False,
                        "message": result.get("error", "Processing failed"),
                        "details": {
                            "errors": result.get("errors", []),
                            "warnings": result.get("warnings", []),
                        },
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        except Exception as e:
            logger.exception("Error in bulk upload")
            return Response(
                {"error": f"Processing failed: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["post"], url_path="reset-all")
    def reset_all_data(self, request):
        """
        Reset all data in the database atomically.

        WARNING: This deletes ALL chapters, members, reports, and analytics data.
        Use with extreme caution!

        Uses transaction.atomic() context manager to:
        - Keep counting logic outside the transaction (read-only, no lock needed)
        - Wrap only the deletion operations in a transaction
        - Ensure ALL deletions succeed or ALL rollback (prevents partial deletion)
        - Maintain database integrity if any deletion fails

        Returns summary of deleted items.
        """
        try:
            from chapters.models import Chapter
            from members.models import Member
            from reports.models import MonthlyReport, MemberMonthlyStats
            from analytics.models import Referral, OneToOne, TYFCB

            # Count before deletion (outside transaction - read-only)
            counts = {
                "chapters": Chapter.objects.count(),
                "members": Member.objects.count(),
                "monthly_reports": MonthlyReport.objects.count(),
                "member_stats": MemberMonthlyStats.objects.count(),
                "referrals": Referral.objects.count(),
                "one_to_ones": OneToOne.objects.count(),
                "tyfcbs": TYFCB.objects.count(),
            }

            # Wrap ALL deletions in atomic transaction
            # If ANY delete fails, ALL are rolled back automatically
            with transaction.atomic():
                # Delete all data (cascade will handle related objects)
                # Order matters: delete dependent objects first to avoid constraint violations
                Chapter.objects.all().delete()
                Member.objects.all().delete()
                MonthlyReport.objects.all().delete()
                MemberMonthlyStats.objects.all().delete()
                Referral.objects.all().delete()
                OneToOne.objects.all().delete()
                TYFCB.objects.all().delete()

            # Transaction completed successfully
            logger.warning(f"Database reset performed. Deleted: {counts}")

            return Response(
                {
                    "success": True,
                    "message": "All data has been deleted successfully",
                    "deleted": counts,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            # Transaction automatically rolled back on exception
            logger.exception("Error resetting database - all changes rolled back")
            return Response(
                {"error": f"Reset failed: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
