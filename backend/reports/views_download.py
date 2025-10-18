"""
MonthlyReport Download ViewSet - Excel and PALMS file download endpoints.

Extracted from views.py to improve maintainability and follow Single Responsibility Principle.
"""

import logging
from io import BytesIO
import openpyxl
from openpyxl.styles import Font

from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response

from chapters.models import Chapter
from chapters.permissions import IsChapterOrAdmin
from reports.models import MonthlyReport

# Import modular formatters
from bni.services.excel_formatters import (
    write_referral_matrix,
    write_oto_matrix,
    write_combination_matrix,
    write_tyfcb_report,
)

logger = logging.getLogger(__name__)


class MonthlyReportDownloadViewSet(viewsets.GenericViewSet):
    """
    ViewSet for MonthlyReport download operations.

    Provides endpoints for:
    - download_matrices: Download all matrices as Excel
    - download_palms: Download original PALMS sheets as ZIP

    Authentication: IsChapterOrAdmin
    - Chapters can access their own reports
    - Admins can access all reports
    """

    queryset = MonthlyReport.objects.all()
    permission_classes = [IsChapterOrAdmin]
    lookup_url_kwarg = 'pk'

    @action(detail=True, methods=["get"], url_path="download-matrices")
    def download_matrices(self, request, pk=None, chapter_id=None):
        """
        Generate and download Excel file with all matrices.

        Creates a workbook with separate sheets for:
        - Referral Matrix
        - One-to-One Matrix
        - Combination Matrix
        - TYFCB Report

        Uses modular formatters for consistent styling across single-month and multi-month reports.
        """
        try:
            chapter = Chapter.objects.get(id=chapter_id)
            monthly_report = MonthlyReport.objects.get(id=pk, chapter=chapter)

            # Import pandas for DataFrame conversion
            import pandas as pd

            # Create a new workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Prepare data for formatters - convert single month to format expected by formatters
            # Period string for single month
            from datetime import datetime
            try:
                month_date = datetime.strptime(monthly_report.month_year, "%Y-%m")
                period_str = month_date.strftime("%m/%Y")
            except:
                period_str = monthly_report.month_year

            # Create a temporary list with this single report for the formatters
            reports_list = [monthly_report]

            # Calculate chapter statistics (needed for performance highlighting)
            # Build aggregated data from this single report
            if monthly_report.referral_matrix_data and "members" in monthly_report.referral_matrix_data:
                members = monthly_report.referral_matrix_data["members"]
                ref_matrix_data = monthly_report.referral_matrix_data["matrix"]

                # Convert to DataFrame
                ref_df = pd.DataFrame(ref_matrix_data, index=members, columns=members)

                # Calculate statistics
                ref_totals = ref_df.sum(axis=1).to_dict()
                avg_referrals = ref_df.sum(axis=1).mean() if len(members) > 0 else 0
            else:
                ref_df = pd.DataFrame()
                ref_totals = {}
                avg_referrals = 0

            if monthly_report.oto_matrix_data and "members" in monthly_report.oto_matrix_data:
                members_oto = monthly_report.oto_matrix_data["members"]
                oto_matrix_data = monthly_report.oto_matrix_data["matrix"]
                oto_df = pd.DataFrame(oto_matrix_data, index=members_oto, columns=members_oto)
                oto_totals = oto_df.sum(axis=1).to_dict()
                avg_oto = oto_df.sum(axis=1).mean() if len(members_oto) > 0 else 0
            else:
                oto_df = pd.DataFrame()
                oto_totals = {}
                avg_oto = 0

            if monthly_report.combination_matrix_data and "members" in monthly_report.combination_matrix_data:
                members_combo = monthly_report.combination_matrix_data["members"]
                combo_matrix_data = monthly_report.combination_matrix_data["matrix"]
                combo_df = pd.DataFrame(combo_matrix_data, index=members_combo, columns=members_combo)
            else:
                combo_df = pd.DataFrame()

            # Calculate TYFCB statistics
            tyfcb_totals = {}
            tyfcb_inside = monthly_report.tyfcb_inside_data or {}
            tyfcb_outside = monthly_report.tyfcb_outside_data or {}

            if tyfcb_inside or tyfcb_outside:
                all_members = set()
                if tyfcb_inside and "by_member" in tyfcb_inside:
                    all_members.update(tyfcb_inside["by_member"].keys())
                if tyfcb_outside and "by_member" in tyfcb_outside:
                    all_members.update(tyfcb_outside["by_member"].keys())

                for member in all_members:
                    inside_amt = float(tyfcb_inside.get("by_member", {}).get(member, 0))
                    outside_amt = float(tyfcb_outside.get("by_member", {}).get(member, 0))
                    tyfcb_totals[member] = inside_amt + outside_amt

            avg_tyfcb = sum(tyfcb_totals.values()) / len(tyfcb_totals) if tyfcb_totals else 0

            # Build stats dict for formatters
            stats = {
                "chapter_size": len(members) if monthly_report.referral_matrix_data else 0,
                "avg_referrals": avg_referrals,
                "avg_oto": avg_oto,
                "avg_tyfcb": avg_tyfcb,
                "ref_totals": ref_totals,
                "oto_totals": oto_totals,
                "tyfcb_totals": tyfcb_totals,
                "total_months": 1,
            }

            # Create sheets using the new formatters
            if monthly_report.referral_matrix_data and not ref_df.empty:
                ws_ref = wb.create_sheet("Referral Matrix")
                write_referral_matrix(ws_ref, ref_df, period_str, stats, reports_list)

            if monthly_report.oto_matrix_data and not oto_df.empty:
                ws_oto = wb.create_sheet("One-to-One Matrix")
                write_oto_matrix(ws_oto, oto_df, period_str, stats, reports_list)

            if monthly_report.combination_matrix_data and not combo_df.empty:
                ws_combo = wb.create_sheet("Combination Matrix")
                write_combination_matrix(ws_combo, combo_df, period_str, stats, reports_list)

            # Create TYFCB sheet using the new formatter
            if tyfcb_inside or tyfcb_outside:
                ws_tyfcb = wb.create_sheet("TYFCB Report")
                # Extract by_member dicts for the formatter
                inside_by_member = tyfcb_inside.get("by_member", {}) if tyfcb_inside else {}
                outside_by_member = tyfcb_outside.get("by_member", {}) if tyfcb_outside else {}
                write_tyfcb_report(ws_tyfcb, inside_by_member, outside_by_member, period_str, stats, reports_list)

            # If no matrices were created, add an info sheet
            if len(wb.sheetnames) == 0:
                ws = wb.create_sheet("Info")
                ws["A1"] = "No matrix data available for this report"
                ws["A1"].font = Font(bold=True)

            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            # Create HTTP response
            filename = f"{chapter.name.replace(' ', '_')}_Matrices_{monthly_report.month_year}.xlsx"
            response = HttpResponse(
                output.getvalue(),  # Use getvalue() instead of read() to avoid buffer position issues
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response

        except Chapter.DoesNotExist:
            return Response(
                {"error": "Chapter not found"}, status=status.HTTP_404_NOT_FOUND
            )
        except MonthlyReport.DoesNotExist:
            return Response(
                {"error": "Monthly report not found"}, status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": f"Failed to generate Excel file: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=True, methods=["get"], url_path="download-palms")
    def download_palms(self, request, pk=None, chapter_id=None):
        """
        Download original PALMS sheets for a report.

        Returns:
        - ZIP file containing all uploaded PALMS sheets if require_palms_sheets is True
        - Error message if PALMS sheets are not available
        """
        import zipfile
        import os
        from django.conf import settings

        try:
            chapter = Chapter.objects.get(id=chapter_id)
            monthly_report = MonthlyReport.objects.get(id=pk, chapter=chapter)

            # Check if PALMS sheets are marked as downloadable
            if not monthly_report.require_palms_sheets:
                return Response(
                    {
                        "error": "PALMS sheets were not marked as downloadable for this report"
                    },
                    status=status.HTTP_404_NOT_FOUND,
                )

            # Check if we have file metadata
            if not monthly_report.uploaded_file_names:
                return Response(
                    {"error": "No file metadata available for this report"},
                    status=status.HTTP_404_NOT_FOUND,
                )

            # Create ZIP file in memory
            zip_buffer = BytesIO()
            files_found = 0

            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                # Add each PALMS file to the ZIP
                for file_info in monthly_report.uploaded_file_names:
                    if file_info.get("file_type") == "slip_audit":
                        # Use saved_filename if available (new format), otherwise fall back to original_filename
                        saved_filename = file_info.get("saved_filename")
                        original_filename = file_info.get(
                            "original_filename", "slip_audit.xls"
                        )
                        file_path_rel = file_info.get(
                            "file_path"
                        )  # e.g., "uploads/filename.xls"

                        media_root = getattr(settings, "MEDIA_ROOT", None)
                        if not media_root:
                            zip_file.writestr(
                                "README.txt",
                                "File storage not configured. Original files may not be available.\n",
                            )
                            continue

                        # Try using the stored file_path first (most reliable)
                        if file_path_rel:
                            file_path = os.path.join(media_root, file_path_rel)
                            if os.path.exists(file_path):
                                with open(file_path, "rb") as f:
                                    zip_file.writestr(original_filename, f.read())
                                files_found += 1
                                continue

                        # Fallback: try using saved_filename
                        if saved_filename:
                            file_path = os.path.join(
                                media_root, "uploads", saved_filename
                            )
                            if os.path.exists(file_path):
                                with open(file_path, "rb") as f:
                                    zip_file.writestr(original_filename, f.read())
                                files_found += 1
                                continue

                        # Last resort: try original filename (backward compatibility for old uploads)
                        file_path = os.path.join(
                            media_root, "uploads", original_filename
                        )
                        if os.path.exists(file_path):
                            with open(file_path, "rb") as f:
                                zip_file.writestr(original_filename, f.read())
                            files_found += 1

                # If no files were found, add an error message
                if files_found == 0:
                    zip_file.writestr(
                        "ERROR.txt",
                        "No PALMS files were found on the server.\n\n"
                        "This could mean:\n"
                        "1. The files were uploaded before file storage was implemented\n"
                        "2. The files were deleted from the server\n"
                        "3. The upload did not complete successfully\n\n"
                        "Please re-upload the report with PALMS sheets enabled.",
                    )

            zip_buffer.seek(0)

            # Generate filename with date
            date_str = (
                monthly_report.week_of_date.strftime("%Y-%m-%d")
                if monthly_report.week_of_date
                else monthly_report.month_year
            )
            filename = f"PALMS_Sheets_{chapter.name.replace(' ', '_')}_{date_str}.zip"

            response = HttpResponse(
                zip_buffer.getvalue(),
                content_type="application/zip",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response

        except Chapter.DoesNotExist:
            return Response(
                {"error": "Chapter not found"}, status=status.HTTP_404_NOT_FOUND
            )
        except MonthlyReport.DoesNotExist:
            return Response(
                {"error": "Monthly report not found"}, status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.exception(f"Failed to download PALMS files for report {pk}")
            return Response(
                {"error": f"Failed to download PALMS files: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
