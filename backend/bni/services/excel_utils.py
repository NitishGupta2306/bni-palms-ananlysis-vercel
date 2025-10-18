"""
Excel formatting utilities for BNI reports.

This module provides reusable Excel styling functions including:
- Merged headers
- Border management
- Separator columns
- Date formatting
"""

from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List


class ExcelFormatter:
    """Utility class for Excel formatting operations."""

    # Color definitions (as per specification.md)
    COLOR_GREEN = "B6FFB6"  # Excellent performance
    COLOR_ORANGE = "FFD699"  # Good/Average performance
    COLOR_RED = "FFB6B6"  # Needs attention
    COLOR_YELLOW = "FFE699"  # Special highlights (non-zero values)
    COLOR_BLACK = "000000"  # Separators
    COLOR_GRAY = "D3D3D3"  # Headers
    COLOR_HEADER_BG = "E8F5E8"  # Soft green for headers

    @staticmethod
    def create_merged_header(
        worksheet, title: str, num_columns: int, row: int = 1, period_str: str = ""
    ):
        """
        Create merged header cell spanning multiple columns.

        Args:
            worksheet: openpyxl worksheet object
            title: Header title text
            num_columns: Number of columns to span
            row: Row number for the header (default: 1)
            period_str: Optional period string to append (e.g., "09/2025 - 11/2025")
        """
        # Build full title
        full_title = f"{title} - Period: {period_str}" if period_str else title

        # Merge cells
        end_col = get_column_letter(num_columns)
        worksheet.merge_cells(f"A{row}:{end_col}{row}")

        # Style the merged cell
        cell = worksheet[f"A{row}"]
        cell.value = full_title
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(
            start_color=ExcelFormatter.COLOR_HEADER_BG,
            end_color=ExcelFormatter.COLOR_HEADER_BG,
            fill_type="solid",
        )

        # Add border
        thick_border = Border(bottom=Side(style="thick"))
        cell.border = thick_border

        # Set fixed row height for merged header (prevents auto-resizing)
        worksheet.row_dimensions[row].height = 30

    @staticmethod
    def get_period_display(reports: List) -> str:
        """
        Get period string in MM/YYYY - MM/YYYY format.

        Args:
            reports: List of MonthlyReport objects (assumed sorted)

        Returns:
            Formatted period string (e.g., "09/2025" or "09/2025 - 11/2025")
        """
        if not reports:
            return ""

        if len(reports) == 1:
            # Single month
            date = datetime.strptime(reports[0].month_year, "%Y-%m")
            return date.strftime("%m/%Y")

        # Multiple months
        start_date = datetime.strptime(reports[0].month_year, "%Y-%m")
        end_date = datetime.strptime(reports[-1].month_year, "%Y-%m")

        return f"{start_date.strftime('%m/%Y')} - {end_date.strftime('%m/%Y')}"

    @staticmethod
    def add_black_separator_column(
        worksheet, col_idx: int, start_row: int, end_row: int
    ):
        """
        Add a black separator column for visual distinction.

        Args:
            worksheet: openpyxl worksheet object
            col_idx: Column index to fill with black
            start_row: Starting row number
            end_row: Ending row number
        """
        col_letter = get_column_letter(col_idx)

        # Fill column with black
        for row_idx in range(start_row, end_row + 1):
            cell = worksheet[f"{col_letter}{row_idx}"]
            cell.fill = PatternFill(
                start_color=ExcelFormatter.COLOR_BLACK,
                end_color=ExcelFormatter.COLOR_BLACK,
                fill_type="solid",
            )

        # Set column width to be narrow (separator effect)
        worksheet.column_dimensions[col_letter].width = 2

    @staticmethod
    def apply_thin_borders(
        worksheet,
        start_row: int,
        end_row: int,
        start_col: int,
        end_col: int,
        skip_columns: List[int] = None,
    ):
        """
        Apply thin borders to a range of cells.

        Args:
            worksheet: openpyxl worksheet object
            start_row: Starting row number
            end_row: Ending row number
            start_col: Starting column index
            end_col: Ending column index
            skip_columns: List of column indices to skip (e.g., separator columns)
        """
        skip_columns = skip_columns or []
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row_idx in range(start_row, end_row + 1):
            for col_idx in range(start_col, end_col + 1):
                if col_idx in skip_columns:
                    continue

                col_letter = get_column_letter(col_idx)
                cell = worksheet[f"{col_letter}{row_idx}"]

                # Preserve existing fill color
                existing_fill = cell.fill
                cell.border = thin_border

                # Reapply fill if it was set (borders can override fills)
                if existing_fill and existing_fill.start_color:
                    cell.fill = existing_fill

    @staticmethod
    def add_thick_right_border(
        worksheet, col_idx: int, start_row: int, end_row: int
    ):
        """
        Add thick border to the right side of a column.

        Args:
            worksheet: openpyxl worksheet object
            col_idx: Column index
            start_row: Starting row number
            end_row: Ending row number
        """
        col_letter = get_column_letter(col_idx)
        thick_side = Side(style="thick")

        for row_idx in range(start_row, end_row + 1):
            cell = worksheet[f"{col_letter}{row_idx}"]
            current_border = cell.border or Border()

            # Create new border preserving existing sides
            cell.border = Border(
                left=current_border.left,
                right=thick_side,
                top=current_border.top,
                bottom=current_border.bottom,
            )

    @staticmethod
    def add_bottom_border_to_row(
        worksheet, row_idx: int, start_col: int, end_col: int, style: str = "thick"
    ):
        """
        Add border to the bottom of a row.

        Args:
            worksheet: openpyxl worksheet object
            row_idx: Row number
            start_col: Starting column index
            end_col: Ending column index
            style: Border style ('thin', 'thick', etc.)
        """
        border_side = Side(style=style)

        for col_idx in range(start_col, end_col + 1):
            col_letter = get_column_letter(col_idx)
            cell = worksheet[f"{col_letter}{row_idx}"]
            current_border = cell.border or Border()

            # Create new border preserving existing sides
            cell.border = Border(
                left=current_border.left,
                right=current_border.right,
                top=current_border.top,
                bottom=border_side,
            )

    @staticmethod
    def add_outer_table_borders(
        worksheet, start_row: int, end_row: int, start_col: int, end_col: int
    ):
        """
        Add thick borders around the outside of a table.

        Args:
            worksheet: openpyxl worksheet object
            start_row: Starting row number
            end_row: Ending row number
            start_col: Starting column index
            end_col: Ending column index
        """
        thick = Side(style="thick")

        # Top border
        for col_idx in range(start_col, end_col + 1):
            col_letter = get_column_letter(col_idx)
            cell = worksheet[f"{col_letter}{start_row}"]
            current_border = cell.border or Border()
            cell.border = Border(
                left=current_border.left,
                right=current_border.right,
                top=thick,
                bottom=current_border.bottom,
            )

        # Bottom border
        for col_idx in range(start_col, end_col + 1):
            col_letter = get_column_letter(col_idx)
            cell = worksheet[f"{col_letter}{end_row}"]
            current_border = cell.border or Border()
            cell.border = Border(
                left=current_border.left,
                right=current_border.right,
                top=current_border.top,
                bottom=thick,
            )

        # Left border
        for row_idx in range(start_row, end_row + 1):
            col_letter = get_column_letter(start_col)
            cell = worksheet[f"{col_letter}{row_idx}"]
            current_border = cell.border or Border()
            cell.border = Border(
                left=thick,
                right=current_border.right,
                top=current_border.top,
                bottom=current_border.bottom,
            )

        # Right border
        for row_idx in range(start_row, end_row + 1):
            col_letter = get_column_letter(end_col)
            cell = worksheet[f"{col_letter}{row_idx}"]
            current_border = cell.border or Border()
            cell.border = Border(
                left=current_border.left,
                right=thick,
                top=current_border.top,
                bottom=current_border.bottom,
            )

    @staticmethod
    def get_month_range(reports: List) -> str:
        """
        Get month range string for display.

        Args:
            reports: List of MonthlyReport objects

        Returns:
            String like "2025-09 to 2025-11" or "2025-09" for single month
        """
        if not reports:
            return ""

        if len(reports) == 1:
            return reports[0].month_year

        return f"{reports[0].month_year} to {reports[-1].month_year}"
