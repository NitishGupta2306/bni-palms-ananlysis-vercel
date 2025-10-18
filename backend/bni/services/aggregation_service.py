"""
Aggregation service for multi-month PALMS analysis.
Combines multiple monthly reports into aggregated matrices and member tracking.
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Set, Tuple
from collections import defaultdict
from datetime import datetime
from io import BytesIO
import zipfile

from reports.models import MonthlyReport
from members.models import Member
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import new modular formatters
from bni.services.excel_formatters import (
    write_referral_matrix,
    write_oto_matrix,
    write_combination_matrix,
    write_tyfcb_report,
    write_summary_page,
    write_inactive_members,
    write_charts_page,
)

# Import refactored helper modules
from bni.services.calculations import PerformanceCalculator
from bni.services.excel_utils import ExcelFormatter
from bni.services.data_aggregator import DataAggregator


class AggregationService:
    """Service for aggregating multiple monthly reports."""

    def __init__(self, reports: List[MonthlyReport]):
        """
        Initialize with list of monthly reports to aggregate.

        Args:
            reports: List of MonthlyReport objects (should be from same chapter)
        """
        self.reports = sorted(reports, key=lambda r: r.month_year)
        self.chapter = self.reports[0].chapter if self.reports else None

        # Color definitions (as per specification.md)
        self.COLOR_GREEN = "B6FFB6"  # Excellent performance
        self.COLOR_ORANGE = "FFD699"  # Good/Average performance
        self.COLOR_RED = "FFB6B6"  # Needs attention
        self.COLOR_YELLOW = "FFE699"  # Special highlights (non-zero values)
        self.COLOR_BLACK = "000000"  # Separators
        self.COLOR_GRAY = "D3D3D3"  # Headers
        self.COLOR_HEADER_BG = "E8F5E8"  # Soft green for headers

        # Performance thresholds (as per specification.md)
        self.THRESHOLD_GREEN = 1.75  # >= 1.75x average
        self.THRESHOLD_ORANGE_HIGH = 1.75  # < 1.75x average
        self.THRESHOLD_ORANGE_LOW = 0.75  # >= 0.75x average
        self.THRESHOLD_RED = 0.5  # < 0.5x average

    # ============================================================================
    # UTILITY METHODS
    # ============================================================================

    def _get_period_display(self) -> str:
        """Get period string in MM/YYYY - MM/YYYY format."""
        return ExcelFormatter.get_period_display(self.reports)

    def _calculate_chapter_statistics(self, aggregated_data: Dict) -> Dict:
        """Calculate chapter-wide statistics for performance evaluation."""
        return PerformanceCalculator.calculate_chapter_statistics(aggregated_data)

    def aggregate_matrices(self) -> Dict:
        """Aggregate all matrices across selected months."""
        return DataAggregator.aggregate_matrices(self.reports, self.chapter)

    def get_member_differences(self) -> List[Dict]:
        """Get list of members who became inactive during the period."""
        return DataAggregator.get_member_differences(self.reports, self.chapter)

    def generate_download_package(self) -> BytesIO:
        """
        Generate ZIP file containing:
        1. Aggregated matrices Excel
        2. Original slip audit files
        3. Member differences report

        Returns:
            BytesIO containing the comprehensive Excel file
        """
        # Generate single comprehensive Excel file
        excel_buffer = self._generate_comprehensive_excel()
        return excel_buffer

    # Data aggregation methods moved to DataAggregator class

    def _generate_comprehensive_excel(self) -> BytesIO:
        """Generate single comprehensive Excel file with all data."""
        aggregated = self.aggregate_matrices()
        differences = self.get_member_differences()

        # Calculate chapter statistics for performance highlighting
        stats = self._calculate_chapter_statistics(aggregated)
        stats["total_months"] = len(self.reports)  # Add total_months to stats

        # Get period string for display
        period_str = self._get_period_display()

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # 1. Summary sheet (first)
        ws_summary = wb.create_sheet("Summary", 0)
        write_summary_page(
            ws_summary,
            self.chapter.name if self.chapter else "BNI",
            period_str,
            aggregated,
            differences,
            stats,
        )

        # 2. Referral Matrix (with monthly breakdowns)
        ws_ref = wb.create_sheet("Referral Matrix")
        write_referral_matrix(
            ws_ref,
            aggregated["referral_matrix"],
            period_str,
            stats,
            self.reports,
        )

        # 3. One-to-One Matrix (with monthly breakdowns)
        ws_oto = wb.create_sheet("One-to-One Matrix")
        write_oto_matrix(
            ws_oto,
            aggregated["oto_matrix"],
            period_str,
            stats,
            self.reports,
        )

        # 4. Combination Matrix (with monthly breakdowns) - uses special 4-column format
        ws_combo = wb.create_sheet("Combination Matrix")
        write_combination_matrix(
            ws_combo,
            aggregated["combination_matrix"],
            period_str,
            stats,
            self.reports,
        )

        # 5. Combined TYFCB Report (Inside and Outside)
        if aggregated["tyfcb_inside"] or aggregated["tyfcb_outside"]:
            ws_tyfcb = wb.create_sheet("TYFCB Report")
            write_tyfcb_report(
                ws_tyfcb,
                aggregated["tyfcb_inside"],
                aggregated["tyfcb_outside"],
                period_str,
                stats,
                self.reports,
            )

        # 6. Member Differences (Inactive Members)
        if differences:
            ws_diff = wb.create_sheet("Inactive Members")
            write_inactive_members(ws_diff, differences, period_str)

        # 7. Charts (Visual Analytics - last sheet)
        ws_charts = wb.create_sheet("Charts")
        write_charts_page(
            ws_charts,
            self.chapter.name if self.chapter else "BNI",
            period_str,
            aggregated,
            stats,
            self.reports,
        )

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer

