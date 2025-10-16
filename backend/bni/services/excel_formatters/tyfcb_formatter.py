"""
TYFCB (Thank You For Closed Business) Report Excel Formatter.

Handles formatting for TYFCB reports with:
- Merged header with period display
- Aggregate table: Inside/Outside/Total TYFCB, Referral counts, Averages
- Monthly breakdown columns for each member
- Special highlighting: RED if Outside > 2x Inside (warning)
- Performance highlighting on Total TYFCB based on chapter average
- Professional borders with section separators
"""

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .colors import (
    COLOR_RED,
    COLOR_GRAY,
    get_performance_color,
)
from .border_utils import (
    create_merged_header,
    apply_standard_table_borders,
)


def write_tyfcb_report(worksheet, tyfcb_inside: dict, tyfcb_outside: dict, period_str: str, stats: dict, reports: list):
    """
    Write TYFCB report with aggregate table and monthly breakdowns.

    Args:
        worksheet: The worksheet to write to
        tyfcb_inside: Dict of {member_name: total_inside_amount}
        tyfcb_outside: Dict of {member_name: total_outside_amount}
        period_str: Period display string (e.g., "01/2025 - 03/2025")
        stats: Chapter statistics dict with avg_tyfcb
        reports: List of MonthlyReport objects for monthly breakdowns
    """
    # Get all unique members
    all_members = set()
    if tyfcb_inside:
        all_members.update(tyfcb_inside.keys())
    if tyfcb_outside:
        all_members.update(tyfcb_outside.keys())

    num_months = len(reports)
    num_agg_cols = 7  # Member, Inside, Outside, Total, Refs, AvgRefs, AvgValue
    total_columns = num_agg_cols + (num_months * 3)  # 3 cols per month

    # Create merged header
    create_merged_header(worksheet, "TYFCB Report", 7, period_str, row=1)

    # Freeze panes
    worksheet.freeze_panes = "B3"

    # =========================================================================
    # BUILD MONTHLY DATA
    # =========================================================================

    monthly_data = {}
    for member in all_members:
        monthly_data[member] = {}
        for idx, report in enumerate(reports, start=1):
            monthly_data[member][idx] = {"inside": 0, "outside": 0, "count": 0}

            # Get inside TYFCB for this month
            if report.tyfcb_inside_data and "by_member" in report.tyfcb_inside_data:
                if member in report.tyfcb_inside_data["by_member"]:
                    monthly_data[member][idx]["inside"] = float(
                        report.tyfcb_inside_data["by_member"][member]
                    )

            # Get outside TYFCB for this month
            if (
                report.tyfcb_outside_data
                and "by_member" in report.tyfcb_outside_data
            ):
                if member in report.tyfcb_outside_data["by_member"]:
                    monthly_data[member][idx]["outside"] = float(
                        report.tyfcb_outside_data["by_member"][member]
                    )

            # Count referrals given this month (from referral matrix)
            if (
                report.referral_matrix_data
                and "members" in report.referral_matrix_data
            ):
                if member in report.referral_matrix_data["members"]:
                    member_idx = report.referral_matrix_data["members"].index(
                        member
                    )
                    if member_idx < len(report.referral_matrix_data["matrix"]):
                        # Count non-zero referrals given by this member
                        row = report.referral_matrix_data["matrix"][member_idx]
                        monthly_data[member][idx]["count"] = sum(
                            1 for val in row if val > 0
                        )

    # =========================================================================
    # COLUMN HEADERS (Row 3, using row 1 for merged header, row 2 blank)
    # =========================================================================

    headers = [
        "Member",
        "Total Inside (AED)",
        "Total Outside (AED)",
        "Total TYFCB (AED)",
        "Total Referrals",
        "Avg Referrals/Month",
        "Avg Value/Referral (AED)",
    ]

    # Add monthly columns
    for idx, report in enumerate(reports, start=1):
        headers.append(f"M{idx} - {report.month_year} Inside")
        headers.append(f"M{idx} - {report.month_year} Outside")
        headers.append(f"M{idx} - {report.month_year} Ref Count")

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type="solid"
        )
        if col_idx > 1:  # Rotate all headers except Member
            cell.alignment = Alignment(textRotation=90, horizontal="center")

    # Set header row height for rotated text visibility
    worksheet.row_dimensions[3].height = 60

    # =========================================================================
    # CALCULATE MEMBER TOTALS
    # =========================================================================

    member_totals = []
    for member in all_members:
        inside_total = tyfcb_inside.get(member, 0) if tyfcb_inside else 0
        outside_total = tyfcb_outside.get(member, 0) if tyfcb_outside else 0
        total_tyfcb = float(inside_total) + float(outside_total)

        # Calculate total referrals across all months
        total_referrals = sum(
            monthly_data[member][idx]["count"]
            for idx in range(1, len(reports) + 1)
        )

        # Calculate averages
        avg_referrals = total_referrals / num_months if num_months > 0 else 0
        avg_value = total_tyfcb / total_referrals if total_referrals > 0 else 0

        member_totals.append(
            {
                "member": member,
                "inside": float(inside_total),
                "outside": float(outside_total),
                "total": total_tyfcb,
                "total_referrals": total_referrals,
                "avg_referrals": avg_referrals,
                "avg_value": avg_value,
            }
        )

    # Sort by total TYFCB descending
    member_totals.sort(key=lambda x: x["total"], reverse=True)

    # Get chapter average TYFCB for performance highlighting
    avg_tyfcb = stats["avg_tyfcb"]

    # =========================================================================
    # DATA ROWS (Starting at row 4)
    # =========================================================================

    row = 4
    for member_data in member_totals:
        member = member_data["member"]
        col = 1

        # Member name - highlight RED if Outside > 2x Inside (warning!)
        member_cell = worksheet.cell(row=row, column=col, value=member)
        member_cell.font = Font(bold=True)

        # Check if Outside TYFCB is more than 2x Inside TYFCB
        inside_val = member_data["inside"]
        outside_val = member_data["outside"]
        if inside_val > 0 and outside_val > (2 * inside_val):
            # Red highlight for name if outside is more than 2x inside
            member_cell.fill = PatternFill(
                start_color=COLOR_RED,
                end_color=COLOR_RED,
                fill_type="solid",
            )
        col += 1

        # Total Inside
        cell = worksheet.cell(row=row, column=col, value=member_data["inside"])
        cell.number_format = "#,##0.00"
        col += 1

        # Total Outside
        cell = worksheet.cell(row=row, column=col, value=member_data["outside"])
        cell.number_format = "#,##0.00"
        col += 1

        # Total TYFCB - with performance highlighting
        cell = worksheet.cell(row=row, column=col, value=member_data["total"])
        cell.number_format = "#,##0.00"
        cell.font = Font(bold=True)

        # Apply performance color based on chapter average
        perf_color = get_performance_color(member_data["total"], avg_tyfcb)
        if perf_color:
            cell.fill = PatternFill(
                start_color=perf_color, end_color=perf_color, fill_type="solid"
            )
        col += 1

        # Total Referrals
        worksheet.cell(row=row, column=col, value=member_data["total_referrals"])
        col += 1

        # Avg Referrals/Month
        cell = worksheet.cell(
            row=row, column=col, value=member_data["avg_referrals"]
        )
        cell.number_format = "0.00"
        col += 1

        # Avg Value/Referral
        cell = worksheet.cell(row=row, column=col, value=member_data["avg_value"])
        cell.number_format = "#,##0.00"
        col += 1

        # Monthly breakdown columns
        for idx in range(1, len(reports) + 1):
            # Inside for this month
            cell = worksheet.cell(
                row=row, column=col, value=monthly_data[member][idx]["inside"]
            )
            cell.number_format = "#,##0.00"
            col += 1

            # Outside for this month
            cell = worksheet.cell(
                row=row, column=col, value=monthly_data[member][idx]["outside"]
            )
            cell.number_format = "#,##0.00"
            col += 1

            # Referral count for this month
            worksheet.cell(
                row=row, column=col, value=monthly_data[member][idx]["count"]
            )
            col += 1

        row += 1

    # =========================================================================
    # TOTALS ROW
    # =========================================================================

    total_row = row + 1
    worksheet.cell(row=total_row, column=1, value="TOTAL:").font = Font(bold=True)

    # Calculate column totals
    total_inside = sum(m["inside"] for m in member_totals)
    total_outside = sum(m["outside"] for m in member_totals)
    total_tyfcb = sum(m["total"] for m in member_totals)
    total_refs = sum(m["total_referrals"] for m in member_totals)

    cell = worksheet.cell(row=total_row, column=2, value=total_inside)
    cell.number_format = "#,##0.00"
    cell.font = Font(bold=True)

    cell = worksheet.cell(row=total_row, column=3, value=total_outside)
    cell.number_format = "#,##0.00"
    cell.font = Font(bold=True)

    cell = worksheet.cell(row=total_row, column=4, value=total_tyfcb)
    cell.number_format = "#,##0.00"
    cell.font = Font(bold=True)

    cell = worksheet.cell(row=total_row, column=5, value=total_refs)
    cell.font = Font(bold=True)

    # =========================================================================
    # BORDERS (Order matters!)
    # =========================================================================

    # Calculate thick separator positions
    thick_separator_cols = [num_agg_cols]  # After "Avg Value/Referral" column
    month_col = num_agg_cols + 1
    for idx in range(num_months - 1):
        month_end = month_col + 2  # After each month's "Ref Count" column
        thick_separator_cols.append(month_end)
        month_col += 3

    # Apply standard borders (outer, header bottom, thick separators, thin grid)
    apply_standard_table_borders(
        worksheet, 3, total_row, 1, total_columns, thick_separator_cols
    )

    # =========================================================================
    # COLUMN WIDTHS
    # =========================================================================

    worksheet.column_dimensions["A"].width = 20  # Member name column
    for col in range(2, total_columns + 1):
        worksheet.column_dimensions[
            get_column_letter(col)
        ].width = 15  # Wider for TYFCB amounts
