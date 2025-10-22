"""
Combination Matrix Excel Formatter.

Handles formatting for combination matrix sheets showing relationship types:
- 0 = Neither (no referral, no OTO)
- 1 = OTO Only
- 2 = Referral Only
- 3 = Both (referral AND OTO)

Features:
- Merged header with period display
- Rotated column headers for member names
- 4 aggregate columns: Both (3), Ref Only (2), OTO Only (1), Neither (0)
- Performance highlighting on "Both" count
- Yellow highlighting on "Both" (value 3) cells
- Monthly breakdown columns (4 per month: Both, Ref, OTO, None)
- Professional borders with section separators
"""

import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .colors import (
    COLOR_YELLOW,
    COLOR_GRAY,
    get_performance_color,
)
from .border_utils import (
    create_merged_header,
    apply_standard_table_borders,
)


def calculate_month_combination(ref_data: dict, oto_data: dict) -> dict:
    """
    Calculate combination matrix for a single month.

    Args:
        ref_data: Referral matrix data for one month
        oto_data: OTO matrix data for one month

    Returns:
        Combined matrix data with values 0-3
    """
    if "members" not in ref_data or "matrix" not in ref_data:
        return None
    if "members" not in oto_data or "matrix" not in oto_data:
        return None

    members = ref_data["members"]
    ref_matrix = ref_data["matrix"]
    oto_matrix = oto_data["matrix"]

    # Create combination matrix: 0=None, 1=OTO, 2=Ref, 3=Both
    combo_matrix = []
    for i in range(len(members)):
        combo_row = []
        for j in range(len(members)):
            ref_val = (
                ref_matrix[i][j]
                if i < len(ref_matrix) and j < len(ref_matrix[i])
                else 0
            )
            oto_val = (
                oto_matrix[i][j]
                if i < len(oto_matrix) and j < len(oto_matrix[i])
                else 0
            )

            combo_val = 0
            if ref_val > 0:
                combo_val += 2
            if oto_val > 0:
                combo_val += 1

            combo_row.append(combo_val)
        combo_matrix.append(combo_row)

    return {"members": members, "matrix": combo_matrix}


def write_combination_matrix(worksheet, aggregated_matrix, period_str: str, stats: dict, reports: list):
    """
    Write combination matrix with 4 aggregate columns and monthly breakdowns.

    Args:
        worksheet: The worksheet to write to
        aggregated_matrix: DataFrame or dict with aggregated combination data
        period_str: Period display string (e.g., "01/2025 - 03/2025")
        stats: Chapter statistics dict (used for "Both" performance highlighting)
        reports: List of MonthlyReport objects for monthly breakdowns
    """
    # Convert to DataFrame if needed
    if isinstance(aggregated_matrix, dict):
        df = pd.DataFrame(aggregated_matrix)
    else:
        df = aggregated_matrix

    num_members = len(df.columns)
    num_months = len(reports)

    # For single-month reports, skip monthly breakdowns (would just duplicate aggregates)
    show_monthly_breakdown = num_months > 1

    # Calculate total columns: 1 (Member col) + num_members + 4 aggregates + (4 monthly cols) * num_months
    if show_monthly_breakdown:
        total_columns = 1 + num_members + 4 + (num_months * 4)
    else:
        total_columns = 1 + num_members + 4  # No monthly columns for single month

    # Create merged header
    create_merged_header(worksheet, "Combination Matrix", total_columns, period_str, row=1)

    # Freeze panes
    worksheet.freeze_panes = "B3"

    # =========================================================================
    # COLUMN HEADERS (Row 2)
    # =========================================================================

    # Member column header
    cell = worksheet.cell(row=2, column=1, value="Member")
    cell.font = Font(bold=True)
    cell.fill = PatternFill(
        start_color=COLOR_GRAY,
        end_color=COLOR_GRAY,
        fill_type="solid",
    )
    cell.alignment = Alignment(horizontal="center", vertical="center")

    current_col = 2

    # Member name column headers (rotated 90°)
    for member_name in df.columns:
        cell = worksheet.cell(row=2, column=current_col)
        cell.value = member_name
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(
            textRotation=90, horizontal="center", vertical="bottom"
        )
        cell.fill = PatternFill(
            start_color=COLOR_GRAY,
            end_color=COLOR_GRAY,
            fill_type="solid",
        )
        current_col += 1

    # Aggregate column headers (4 columns for combination)
    agg_start_col = current_col
    agg_headers = ["Both (3)", "Ref Only (2)", "OTO Only (1)", "Neither (0)"]
    for header in agg_headers:
        cell = worksheet.cell(row=2, column=current_col, value=header)
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill(
            start_color=COLOR_GRAY,
            end_color=COLOR_GRAY,
            fill_type="solid",
        )
        cell.alignment = Alignment(wrap_text=True, horizontal="center")
        current_col += 1

    agg_end_col = current_col - 1

    # Monthly column headers (4 columns per month for combination, only for multi-month reports)
    if show_monthly_breakdown:
        for idx, report in enumerate(reports, start=1):
            month_date = datetime.strptime(report.month_year, "%Y-%m")
            month_display = month_date.strftime("%m/%Y")

            for combo_label in ["Both", "Ref", "OTO", "None"]:
                cell = worksheet.cell(
                    row=2,
                    column=current_col,
                    value=f"M{idx}-{month_display}\n{combo_label}",
                )
                cell.font = Font(bold=True, size=7)
                cell.alignment = Alignment(
                    textRotation=90,
                    horizontal="center",
                    vertical="bottom",
                    wrap_text=True,
                )
                cell.fill = PatternFill(
                    start_color=COLOR_GRAY,
                    end_color=COLOR_GRAY,
                    fill_type="solid",
                )
                current_col += 1

    # Set header row height for rotated text visibility
    worksheet.row_dimensions[2].height = 60

    # =========================================================================
    # DATA ROWS (Starting at row 3)
    # =========================================================================

    # Calculate "Both" counts for performance highlighting
    member_both_counts = {}
    for member in df.index:
        member_both_counts[member] = (df.loc[member] == 3).sum()
    avg_both = (
        sum(member_both_counts.values()) / len(member_both_counts)
        if member_both_counts
        else 0
    )

    yellow_fill = PatternFill(
        start_color=COLOR_YELLOW,
        end_color=COLOR_YELLOW,
        fill_type="solid",
    )

    current_row = 3
    for row_name in df.index:
        # Member name (with performance highlighting based on "Both" count)
        member_cell = worksheet.cell(row=current_row, column=1, value=row_name)
        member_cell.font = Font(bold=True)

        perf_color = get_performance_color(
            member_both_counts[row_name], avg_both
        )
        if perf_color:
            member_cell.fill = PatternFill(
                start_color=perf_color, end_color=perf_color, fill_type="solid"
            )

        # Matrix data cells
        col_idx = 2
        for col_name in df.columns:
            value = df.loc[row_name, col_name]
            cell = worksheet.cell(row=current_row, column=col_idx, value=value)

            # Yellow highlight for "Both" (value 3)
            if value == 3:
                cell.fill = yellow_fill
                cell.font = Font(bold=True)

            col_idx += 1

        # Calculate aggregate counts for this member
        member_row = df.loc[row_name]
        both_count = (member_row == 3).sum()
        ref_only_count = (member_row == 2).sum()
        oto_only_count = (member_row == 1).sum()
        neither_count = (member_row == 0).sum()

        # Write aggregate columns (with performance highlighting on "Both" only)
        agg_values = [both_count, ref_only_count, oto_only_count, neither_count]
        for i, agg_val in enumerate(agg_values):
            cell = worksheet.cell(
                row=current_row, column=agg_start_col + i, value=agg_val
            )
            cell.font = Font(bold=True)
            if i == 0 and perf_color:  # Only highlight "Both" column
                cell.fill = PatternFill(
                    start_color=perf_color, end_color=perf_color, fill_type="solid"
                )

        # Monthly counts (only for multi-month reports)
        if show_monthly_breakdown:
            # Move to monthly columns (right after 4 aggregate columns)
            col_idx = agg_end_col + 1

            for report in reports:
                # Get month combination data
                if report.referral_matrix_data and report.oto_matrix_data:
                    month_matrix_data = calculate_month_combination(
                        report.referral_matrix_data, report.oto_matrix_data
                    )
                else:
                    month_matrix_data = None

                # Calculate monthly counts for this member
                if (
                    month_matrix_data
                    and "members" in month_matrix_data
                    and "matrix" in month_matrix_data
                ):
                    members = month_matrix_data["members"]
                    matrix = month_matrix_data["matrix"]

                    if row_name in members:
                        member_idx = members.index(row_name)
                        if member_idx < len(matrix):
                            month_row = matrix[member_idx]
                            month_both = sum(1 for val in month_row if val == 3)
                            month_ref = sum(1 for val in month_row if val == 2)
                            month_oto = sum(1 for val in month_row if val == 1)
                            month_none = sum(1 for val in month_row if val == 0)

                            worksheet.cell(
                                row=current_row, column=col_idx, value=month_both
                            )
                            worksheet.cell(
                                row=current_row, column=col_idx + 1, value=month_ref
                            )
                            worksheet.cell(
                                row=current_row, column=col_idx + 2, value=month_oto
                            )
                            worksheet.cell(
                                row=current_row, column=col_idx + 3, value=month_none
                            )
                        else:
                            for i in range(4):
                                worksheet.cell(
                                    row=current_row, column=col_idx + i, value=0
                                )
                    else:
                        for i in range(4):
                            worksheet.cell(row=current_row, column=col_idx + i, value=0)
                else:
                    for i in range(4):
                        worksheet.cell(row=current_row, column=col_idx + i, value=0)

                col_idx += 4  # Move to next month

        current_row += 1

    # =========================================================================
    # TOTAL RECEIVED ROW
    # =========================================================================

    total_row = current_row
    worksheet.cell(row=total_row, column=1, value="Total Received").font = Font(
        bold=True
    )

    # Totals for member name columns (matrix cells) - show "Both" count
    col_idx = 2
    for col_name in df.columns:
        col_both_count = (df[col_name] == 3).sum()
        cell = worksheet.cell(row=total_row, column=col_idx, value=col_both_count)
        cell.font = Font(bold=True)
        col_idx += 1

    # Totals for aggregate columns (sum all members' aggregate values)
    for agg_idx in range(4):  # Both, Ref Only, OTO Only, Neither
        col_total = 0
        for row_name in df.index:
            member_row = df.loc[row_name]
            if agg_idx == 0:  # Both
                col_total += (member_row == 3).sum()
            elif agg_idx == 1:  # Ref Only
                col_total += (member_row == 2).sum()
            elif agg_idx == 2:  # OTO Only
                col_total += (member_row == 1).sum()
            else:  # Neither
                col_total += (member_row == 0).sum()

        cell = worksheet.cell(
            row=total_row, column=agg_start_col + agg_idx, value=col_total
        )
        cell.font = Font(bold=True)

    # =========================================================================
    # BORDERS (Order matters!)
    # =========================================================================

    # Calculate thick separator positions
    thick_separator_cols = [agg_end_col]  # After "Neither (0)" column

    # Add monthly separators only for multi-month reports
    if show_monthly_breakdown:
        month_col = agg_end_col + 1
        for idx in range(num_months - 1):
            month_end = month_col + 3  # After each month's "None" column
            thick_separator_cols.append(month_end)
            month_col += 4

    # Apply standard borders (outer, header bottom, thick separators, thin grid)
    apply_standard_table_borders(
        worksheet, 2, total_row, 1, total_columns, thick_separator_cols
    )

    # =========================================================================
    # COLUMN WIDTHS
    # =========================================================================

    # Member name column (first column)
    worksheet.column_dimensions["A"].width = 20

    # Member name columns (matrix columns - narrower for rotated text)
    for col in range(2, 2 + num_members):
        worksheet.column_dimensions[get_column_letter(col)].width = 4

    # Aggregate columns (Neither, OTO only, etc.)
    for col in range(agg_start_col, agg_end_col + 1):
        worksheet.column_dimensions[get_column_letter(col)].width = 15

    # Monthly columns
    if show_monthly_breakdown:
        for col in range(agg_end_col + 1, total_columns + 1):
            worksheet.column_dimensions[get_column_letter(col)].width = 10
