"""
Charts Excel Formatter.

Creates visual charts for:
- Performance distribution pie chart
- Top performers bar chart
- Monthly trends line chart
- Inside vs Outside TYFCB comparison
"""

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList

from .colors import COLOR_GREEN, COLOR_ORANGE, COLOR_RED, COLOR_GRAY


def write_charts_page(worksheet, chapter_name: str, period_str: str, aggregated_data: dict,
                     stats: dict, reports: list):
    """
    Write charts page with visual insights.

    Args:
        worksheet: The worksheet to write to
        chapter_name: Name of the chapter
        period_str: Period display string
        aggregated_data: Dict with referral_matrix, oto_matrix, etc.
        stats: Chapter statistics dict
        reports: List of MonthlyReport objects
    """
    # Set print settings
    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.paperSize = 1  # Letter

    # Title
    worksheet.merge_cells('A1:M1')
    title_cell = worksheet['A1']
    title_cell.value = f"{chapter_name} - Visual Analytics"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    worksheet.row_dimensions[1].height = 30

    worksheet.merge_cells('A2:M2')
    period_cell = worksheet['A2']
    period_cell.value = f"Period: {period_str}"
    period_cell.font = Font(size=11, italic=True)
    period_cell.alignment = Alignment(horizontal='center')

    # =========================================================================
    # CHART 1: PERFORMANCE DISTRIBUTION PIE CHART
    # =========================================================================

    from .colors import count_performance_tiers
    ref_tiers = count_performance_tiers(stats["ref_totals"], stats["avg_referrals"])

    # Data for pie chart
    worksheet['A5'] = 'Performance Tier'
    worksheet['B5'] = 'Member Count'
    worksheet['A5'].font = Font(bold=True)
    worksheet['B5'].font = Font(bold=True)

    worksheet['A6'] = 'Excellent (Green)'
    worksheet['B6'] = ref_tiers.get('green', 0)
    worksheet['A7'] = 'Good/Average (Orange)'
    worksheet['B7'] = ref_tiers.get('orange', 0)
    worksheet['A8'] = 'Needs Attention (Red)'
    worksheet['B8'] = ref_tiers.get('red', 0)

    # Create pie chart
    pie = PieChart()
    labels = Reference(worksheet, min_col=1, min_row=6, max_row=8)
    data = Reference(worksheet, min_col=2, min_row=5, max_row=8)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Performance Distribution"
    pie.height = 10
    pie.width = 15

    # Add data labels showing percentages
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True

    worksheet.add_chart(pie, "D5")

    # =========================================================================
    # CHART 2: TOP 10 PERFORMERS BAR CHART
    # =========================================================================

    # Get top 10 performers
    ref_matrix = aggregated_data["referral_matrix"]
    member_performance = []

    for member in ref_matrix.index:
        ref_val = stats["ref_totals"].get(member, 0)
        oto_val = stats["oto_totals"].get(member, 0)
        tyfcb_val = stats["tyfcb_totals"].get(member, 0)

        member_performance.append({
            "member": member,
            "ref": ref_val,
            "oto": oto_val,
            "tyfcb": tyfcb_val,
        })

    # Sort by referrals
    member_performance.sort(key=lambda x: x["ref"], reverse=True)
    top_10 = member_performance[:10]

    # Data for bar chart
    start_row = 20
    worksheet.cell(row=start_row, column=1, value='Member').font = Font(bold=True)
    worksheet.cell(row=start_row, column=2, value='Referrals').font = Font(bold=True)
    worksheet.cell(row=start_row, column=3, value='OTO').font = Font(bold=True)
    worksheet.cell(row=start_row, column=4, value='TYFCB (AED)').font = Font(bold=True)

    for i, perf in enumerate(top_10):
        row = start_row + 1 + i
        worksheet.cell(row=row, column=1, value=perf["member"])
        worksheet.cell(row=row, column=2, value=perf["ref"])
        worksheet.cell(row=row, column=3, value=perf["oto"])
        worksheet.cell(row=row, column=4, value=perf["tyfcb"])

    # Create bar chart
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Top 10 Performers - Referrals, OTO, TYFCB"
    bar.y_axis.title = 'Count / Amount'
    bar.x_axis.title = 'Members'

    data = Reference(worksheet, min_col=2, min_row=start_row, max_row=start_row + 10, max_col=4)
    cats = Reference(worksheet, min_col=1, min_row=start_row + 1, max_row=start_row + 10)

    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.height = 12
    bar.width = 20

    worksheet.add_chart(bar, "F20")

    # =========================================================================
    # CHART 3: MONTHLY TRENDS LINE CHART
    # =========================================================================

    if len(reports) > 1:
        # Calculate monthly totals
        start_row = 35
        worksheet.cell(row=start_row, column=1, value='Month').font = Font(bold=True)
        worksheet.cell(row=start_row, column=2, value='Total Referrals').font = Font(bold=True)
        worksheet.cell(row=start_row, column=3, value='Total OTO').font = Font(bold=True)
        worksheet.cell(row=start_row, column=4, value='Total TYFCB (AED)').font = Font(bold=True)

        for i, report in enumerate(reports):
            row = start_row + 1 + i

            # Month label
            try:
                from datetime import datetime
                month_date = datetime.strptime(report.month_year, "%Y-%m")
                month_label = month_date.strftime("%b %Y")
            except:
                month_label = report.month_year

            worksheet.cell(row=row, column=1, value=month_label)

            # Calculate totals for this month
            if report.referral_matrix_data and "matrix" in report.referral_matrix_data:
                ref_total = sum(sum(row) for row in report.referral_matrix_data["matrix"])
                worksheet.cell(row=row, column=2, value=ref_total)

            if report.oto_matrix_data and "matrix" in report.oto_matrix_data:
                oto_total = sum(sum(row) for row in report.oto_matrix_data["matrix"])
                worksheet.cell(row=row, column=3, value=oto_total)

            # TYFCB total
            tyfcb_total = 0
            if report.tyfcb_inside_data:
                tyfcb_total += float(report.tyfcb_inside_data.get('total_amount', 0))
            if report.tyfcb_outside_data:
                tyfcb_total += float(report.tyfcb_outside_data.get('total_amount', 0))
            worksheet.cell(row=row, column=4, value=tyfcb_total)

        # Create line chart
        line = LineChart()
        line.title = "Monthly Trends"
        line.style = 12
        line.y_axis.title = 'Count / Amount (AED)'
        line.x_axis.title = 'Month'

        data = Reference(worksheet, min_col=2, min_row=start_row, max_row=start_row + len(reports), max_col=4)
        cats = Reference(worksheet, min_col=1, min_row=start_row + 1, max_row=start_row + len(reports))

        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        line.height = 10
        line.width = 18

        worksheet.add_chart(line, "A40")

    # =========================================================================
    # CHART 4: INSIDE VS OUTSIDE TYFCB BAR CHART
    # =========================================================================

    if stats.get("tyfcb_totals"):
        # Get top 10 by TYFCB
        tyfcb_performance = []
        for member, total in stats["tyfcb_totals"].items():
            if total > 0:
                # Get inside and outside breakdown
                inside = 0
                outside = 0
                if aggregated_data.get("tyfcb_inside"):
                    inside = float(aggregated_data["tyfcb_inside"].get(member, 0))
                if aggregated_data.get("tyfcb_outside"):
                    outside = float(aggregated_data["tyfcb_outside"].get(member, 0))

                tyfcb_performance.append({
                    "member": member,
                    "inside": inside,
                    "outside": outside,
                    "total": total
                })

        tyfcb_performance.sort(key=lambda x: x["total"], reverse=True)
        top_10_tyfcb = tyfcb_performance[:10]

        # Data for stacked bar chart
        start_row = 55
        worksheet.cell(row=start_row, column=1, value='Member').font = Font(bold=True)
        worksheet.cell(row=start_row, column=2, value='Inside Chapter (AED)').font = Font(bold=True)
        worksheet.cell(row=start_row, column=3, value='Outside Chapter (AED)').font = Font(bold=True)

        for i, perf in enumerate(top_10_tyfcb):
            row = start_row + 1 + i
            worksheet.cell(row=row, column=1, value=perf["member"])
            worksheet.cell(row=row, column=2, value=perf["inside"])
            worksheet.cell(row=row, column=3, value=perf["outside"])

        # Create stacked bar chart
        bar2 = BarChart()
        bar2.type = "col"
        bar2.grouping = "stacked"
        bar2.overlap = 100
        bar2.title = "Top 10 TYFCB - Inside vs Outside Chapter"
        bar2.y_axis.title = 'Amount (AED)'
        bar2.x_axis.title = 'Members'

        data = Reference(worksheet, min_col=2, min_row=start_row, max_row=start_row + len(top_10_tyfcb), max_col=3)
        cats = Reference(worksheet, min_col=1, min_row=start_row + 1, max_row=start_row + len(top_10_tyfcb))

        bar2.add_data(data, titles_from_data=True)
        bar2.set_categories(cats)
        bar2.height = 12
        bar2.width = 20

        worksheet.add_chart(bar2, "F55")

    # =========================================================================
    # COLUMN WIDTHS
    # =========================================================================

    worksheet.column_dimensions['A'].width = 25
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 15
    worksheet.column_dimensions['D'].width = 15
