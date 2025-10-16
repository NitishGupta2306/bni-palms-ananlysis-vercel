"""
Executive Summary Excel Formatter.

Creates a one-page executive summary optimized for presentations with:
- Large key metrics
- Top performers highlights
- Areas needing attention
- Performance distribution pie chart
- Trends summary
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference
from datetime import datetime

from .colors import (
    COLOR_GREEN,
    COLOR_ORANGE,
    COLOR_RED,
    COLOR_GRAY,
    COLOR_HEADER_BG,
)


def write_executive_summary(worksheet, chapter_name: str, period_str: str, aggregated_data: dict,
                            differences: list, stats: dict):
    """
    Write executive summary page optimized for presentations.

    Args:
        worksheet: The worksheet to write to
        chapter_name: Name of the chapter
        period_str: Period display string (e.g., "01/2025 - 03/2025")
        aggregated_data: Dict with referral_matrix, oto_matrix, etc.
        differences: List of inactive members
        stats: Chapter statistics dict
    """
    # Set print settings
    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.paperSize = 1  # Letter
    worksheet.page_setup.fitToPage = True
    worksheet.page_setup.fitToHeight = 1
    worksheet.page_setup.fitToWidth = 1

    # =========================================================================
    # HEADER SECTION
    # =========================================================================

    # Main title
    worksheet.merge_cells('A1:H1')
    title_cell = worksheet['A1']
    title_cell.value = f"{chapter_name} - Executive Summary"
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(
        start_color=COLOR_HEADER_BG,
        end_color=COLOR_HEADER_BG,
        fill_type='solid'
    )
    worksheet.row_dimensions[1].height = 35

    # Period subtitle
    worksheet.merge_cells('A2:H2')
    period_cell = worksheet['A2']
    period_cell.value = f"Period: {period_str}"
    period_cell.font = Font(size=12, italic=True)
    period_cell.alignment = Alignment(horizontal='center')
    worksheet.row_dimensions[2].height = 20

    # Generation timestamp
    worksheet.merge_cells('A3:H3')
    gen_cell = worksheet['A3']
    gen_cell.value = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    gen_cell.font = Font(size=9, color='666666')
    gen_cell.alignment = Alignment(horizontal='center')

    current_row = 5

    # =========================================================================
    # KEY METRICS SECTION (Large, Bold)
    # =========================================================================

    worksheet.merge_cells(f'A{current_row}:H{current_row}')
    section_cell = worksheet.cell(row=current_row, column=1)
    section_cell.value = "KEY METRICS"
    section_cell.font = Font(bold=True, size=14)
    section_cell.fill = PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type='solid')
    section_cell.alignment = Alignment(horizontal='center')
    current_row += 1

    # Key metrics in large boxes
    metrics = [
        ("Chapter Size", stats["chapter_size"], "members"),
        ("Avg Referrals", f"{stats['avg_referrals']:.1f}", "per member"),
        ("Avg OTO", f"{stats['avg_oto']:.1f}", "per member"),
        ("Avg TYFCB", f"AED {stats['avg_tyfcb']:,.0f}", "per member"),
    ]

    for i, (label, value, unit) in enumerate(metrics):
        col = 1 + (i * 2)

        # Merge 2 columns for each metric
        worksheet.merge_cells(start_row=current_row, start_column=col,
                            end_row=current_row + 2, end_column=col + 1)

        metric_cell = worksheet.cell(row=current_row, column=col)
        metric_cell.value = f"{value}\n{label}\n({unit})"
        metric_cell.font = Font(bold=True, size=16)
        metric_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        metric_cell.fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')

        # Add border
        border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
        metric_cell.border = border

    worksheet.row_dimensions[current_row].height = 25
    worksheet.row_dimensions[current_row + 1].height = 25
    worksheet.row_dimensions[current_row + 2].height = 25
    current_row += 4

    # =========================================================================
    # PERFORMANCE DISTRIBUTION
    # =========================================================================

    worksheet.merge_cells(f'A{current_row}:D{current_row}')
    section_cell = worksheet.cell(row=current_row, column=1)
    section_cell.value = "PERFORMANCE DISTRIBUTION"
    section_cell.font = Font(bold=True, size=12)
    section_cell.fill = PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type='solid')
    current_row += 1

    # Calculate performance tiers
    from .colors import count_performance_tiers
    ref_tiers = count_performance_tiers(stats["ref_totals"], stats["avg_referrals"])

    # Performance distribution table
    perf_data = [
        ("🟢 Excellent", f"{ref_tiers['green_pct']:.0f}%", ref_tiers['green_count'], COLOR_GREEN),
        ("🟠 Good/Average", f"{ref_tiers['orange_pct']:.0f}%", ref_tiers['orange_count'], COLOR_ORANGE),
        ("🔴 Needs Attention", f"{ref_tiers['red_pct']:.0f}%", ref_tiers['red_count'], COLOR_RED),
    ]

    for label, pct, count, color in perf_data:
        worksheet.cell(row=current_row, column=1, value=label).font = Font(bold=True)
        worksheet.cell(row=current_row, column=2, value=pct).font = Font(size=14, bold=True)
        worksheet.cell(row=current_row, column=3, value=f"({count} members)")

        # Color the percentage cell
        pct_cell = worksheet.cell(row=current_row, column=2)
        pct_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        pct_cell.alignment = Alignment(horizontal='center')

        current_row += 1

    current_row += 1

    # =========================================================================
    # TOP PERFORMERS
    # =========================================================================

    worksheet.merge_cells(f'E{current_row - 4}:H{current_row - 4}')
    section_cell = worksheet.cell(row=current_row - 4, column=5)
    section_cell.value = "⭐ TOP PERFORMERS"
    section_cell.font = Font(bold=True, size=12)
    section_cell.fill = PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type='solid')

    # Get top 5 performers
    ref_matrix = aggregated_data["referral_matrix"]
    member_performance = []

    for member in ref_matrix.index:
        ref_val = stats["ref_totals"].get(member, 0)
        oto_val = stats["oto_totals"].get(member, 0)
        tyfcb_val = stats["tyfcb_totals"].get(member, 0)

        # Calculate normalized score
        ref_ratio = ref_val / stats["avg_referrals"] if stats["avg_referrals"] > 0 else 0
        oto_ratio = oto_val / stats["avg_oto"] if stats["avg_oto"] > 0 else 0
        tyfcb_ratio = tyfcb_val / stats["avg_tyfcb"] if stats["avg_tyfcb"] > 0 else 0
        total_score = ref_ratio + oto_ratio + tyfcb_ratio

        member_performance.append({
            "member": member,
            "score": total_score,
            "ref": ref_val,
            "oto": oto_val,
            "tyfcb": tyfcb_val,
        })

    # Sort and get top 5
    member_performance.sort(key=lambda x: x["score"], reverse=True)
    top_5 = member_performance[:5]

    top_row = current_row - 3
    for rank, perf in enumerate(top_5, start=1):
        worksheet.cell(row=top_row, column=5, value=f"#{rank}").font = Font(bold=True, size=12)
        worksheet.cell(row=top_row, column=6, value=perf["member"]).font = Font(bold=True)

        # Show their best metric
        best_metric = "Referrals"
        best_value = perf["ref"]
        if perf["tyfcb"] / stats["avg_tyfcb"] > perf["ref"] / stats["avg_referrals"]:
            best_metric = "TYFCB"
            best_value = f"AED {perf['tyfcb']:,.0f}"

        worksheet.cell(row=top_row, column=7, value=best_metric).font = Font(size=9)
        worksheet.cell(row=top_row, column=8, value=best_value).font = Font(size=9)

        # Green background for top performer
        if rank == 1:
            for col in range(5, 9):
                worksheet.cell(row=top_row, column=col).fill = PatternFill(
                    start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type='solid'
                )

        top_row += 1

    # =========================================================================
    # AREAS NEEDING ATTENTION
    # =========================================================================

    worksheet.merge_cells(f'A{current_row}:D{current_row}')
    section_cell = worksheet.cell(row=current_row, column=1)
    section_cell.value = "⚠️  AREAS NEEDING ATTENTION"
    section_cell.font = Font(bold=True, size=12)
    section_cell.fill = PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type='solid')
    current_row += 1

    # Get bottom 3 performers
    bottom_3 = member_performance[-3:]
    bottom_3.reverse()  # Show worst first

    for perf in bottom_3:
        worksheet.cell(row=current_row, column=1, value=perf["member"]).font = Font(bold=True)
        worksheet.cell(row=current_row, column=2, value=f"Refs: {perf['ref']}")
        worksheet.cell(row=current_row, column=3, value=f"OTO: {perf['oto']}")
        worksheet.cell(row=current_row, column=4, value=f"TYFCB: AED {perf['tyfcb']:,.0f}")

        # Red background for low performers
        for col in range(1, 5):
            worksheet.cell(row=current_row, column=col).fill = PatternFill(
                start_color=COLOR_RED, end_color=COLOR_RED, fill_type='solid'
            )

        current_row += 1

    current_row += 1

    # =========================================================================
    # INACTIVE MEMBERS WARNING
    # =========================================================================

    if differences:
        worksheet.merge_cells(f'E{current_row - 3}:H{current_row - 3}')
        section_cell = worksheet.cell(row=current_row - 3, column=5)
        section_cell.value = f"🚫 {len(differences)} INACTIVE MEMBERS"
        section_cell.font = Font(bold=True, size=12, color='FFFFFF')
        section_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        section_cell.alignment = Alignment(horizontal='center')

        inactive_row = current_row - 2
        for member_data in differences[:3]:  # Show up to 3
            worksheet.cell(row=inactive_row, column=5, value=member_data["member_name"])
            worksheet.cell(row=inactive_row, column=6, value=member_data["last_active_month"]).font = Font(size=9)
            inactive_row += 1

    # =========================================================================
    # SUMMARY INSIGHTS
    # =========================================================================

    current_row += 1
    worksheet.merge_cells(f'A{current_row}:H{current_row}')
    section_cell = worksheet.cell(row=current_row, column=1)
    section_cell.value = "📊 SUMMARY INSIGHTS"
    section_cell.font = Font(bold=True, size=12)
    section_cell.fill = PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type='solid')
    current_row += 1

    # Generate insights
    insights = []

    if ref_tiers['green_pct'] > 30:
        insights.append(f"✓ Strong performance: {ref_tiers['green_pct']:.0f}% of members are excellent performers")

    if ref_tiers['red_pct'] > 30:
        insights.append(f"⚠ {ref_tiers['red_pct']:.0f}% of members need coaching and support")

    if differences:
        insights.append(f"⚠ {len(differences)} members became inactive during this period")

    if stats['avg_tyfcb'] > 1000:
        insights.append(f"✓ High TYFCB average: AED {stats['avg_tyfcb']:,.0f} per member")

    if not insights:
        insights.append("Chapter performance is balanced across all tiers")

    for insight in insights:
        worksheet.cell(row=current_row, column=1, value=insight).font = Font(size=11)
        worksheet.merge_cells(f'A{current_row}:H{current_row}')
        current_row += 1

    # =========================================================================
    # COLUMN WIDTHS
    # =========================================================================

    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 15
    worksheet.column_dimensions['D'].width = 15
    worksheet.column_dimensions['E'].width = 15
    worksheet.column_dimensions['F'].width = 20
    worksheet.column_dimensions['G'].width = 15
    worksheet.column_dimensions['H'].width = 15

    # Add footer
    footer_row = current_row + 2
    worksheet.merge_cells(f'A{footer_row}:H{footer_row}')
    footer_cell = worksheet.cell(row=footer_row, column=1)
    footer_cell.value = f"BNI PALMS Analytics | {chapter_name} | Report v1.0"
    footer_cell.font = Font(size=8, color='999999', italic=True)
    footer_cell.alignment = Alignment(horizontal='center')
