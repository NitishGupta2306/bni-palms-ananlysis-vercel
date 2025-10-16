"""
Enhanced Summary Page Excel Formatter.

Combines the best of Executive Summary and Summary pages with:
- Large, bold key metrics boxes for quick overview
- Chapter statistics with contextual insights (on the right)
- Top 5 Performers section with growth percentages
- Bottom 3 Need Attention section (RED highlighted) with drop percentages
- Hyperlink button to jump to full All Members table
- Performance Guide (color legend) positioned on the right side
- Full All Members performance table (positioned 35 rows below)
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

from .colors import (
    COLOR_GREEN,
    COLOR_ORANGE,
    COLOR_RED,
    COLOR_GRAY,
    COLOR_HEADER_BG,
    get_performance_color,
    count_performance_tiers,
)
from .border_utils import create_merged_header, configure_print_settings


def write_summary_page(worksheet, chapter_name: str, period_str: str, aggregated_data: dict,
                       differences: list, stats: dict):
    """
    Write enhanced summary page with large metrics, insights, and full member data.

    Args:
        worksheet: The worksheet to write to
        chapter_name: Name of the chapter
        period_str: Period display string (e.g., "01/2025 - 03/2025")
        aggregated_data: Dict with referral_matrix, oto_matrix, etc.
        differences: List of inactive members
        stats: Chapter statistics dict
    """
    # Configure print settings
    configure_print_settings(worksheet, orientation='landscape', fit_to_page=True)

    # Merged header
    create_merged_header(
        worksheet,
        f"{chapter_name} - Summary Report",
        12,
        period_str,
        row=1,
    )

    # Freeze panes
    worksheet.freeze_panes = "A3"

    current_row = 3

    # =========================================================================
    # SECTION 1: LARGE KEY METRICS BOXES (LEFT SIDE - 4 boxes)
    # =========================================================================

    # Section header
    worksheet.merge_cells(f'A{current_row}:H{current_row}')
    section_cell = worksheet.cell(row=current_row, column=1)
    section_cell.value = "KEY METRICS"
    section_cell.font = Font(bold=True, size=14)
    section_cell.fill = PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type='solid')
    section_cell.alignment = Alignment(horizontal='center')
    current_row += 1

    # Large metric boxes (4 boxes, each spans 2 columns and 3 rows)
    metrics = [
        ("Chapter Size", stats["chapter_size"], "members"),
        ("Avg Referrals", f"{stats['avg_referrals']:.1f}", "per member"),
        ("Avg OTO", f"{stats['avg_oto']:.1f}", "per member"),
        ("Avg TYFCB", f"AED {stats['avg_tyfcb']:,.0f}", "per member"),
    ]

    for i, (label, value, unit) in enumerate(metrics):
        col = 1 + (i * 2)

        # Merge 2 columns x 3 rows for each metric
        worksheet.merge_cells(start_row=current_row, start_column=col,
                            end_row=current_row + 2, end_column=col + 1)

        metric_cell = worksheet.cell(row=current_row, column=col)
        metric_cell.value = f"{value}\n{label}\n({unit})"
        metric_cell.font = Font(bold=True, size=16)
        metric_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        metric_cell.fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type='solid')

        # Add border
        border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
        metric_cell.border = border

    worksheet.row_dimensions[current_row].height = 25
    worksheet.row_dimensions[current_row + 1].height = 25
    worksheet.row_dimensions[current_row + 2].height = 25
    current_row += 4

    # =========================================================================
    # SECTION 2: PERFORMANCE GUIDE (TOP RIGHT - POSITIONED NEXT TO METRICS)
    # =========================================================================

    guide_start_row = 4  # Aligns with metrics section
    guide_col = 10

    # Guide header
    worksheet.merge_cells(
        start_row=guide_start_row,
        start_column=guide_col,
        end_row=guide_start_row,
        end_column=guide_col + 2,
    )
    guide_header_cell = worksheet.cell(row=guide_start_row, column=guide_col)
    guide_header_cell.value = "Performance Guide"
    guide_header_cell.font = Font(bold=True, size=12)
    guide_header_cell.fill = PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type='solid')
    guide_header_cell.alignment = Alignment(horizontal='center')
    guide_start_row += 1

    # Guide table
    guide_data = [
        ("Color", "Meaning", "Threshold"),
        ("Green", "Excellent", "≥ 1.75x average"),
        ("Orange", "Good/Average", "0.75x - 1.75x average"),
        ("Red", "Needs Attention", "< 0.5x average"),
    ]

    for row_offset, (color_name, meaning, threshold) in enumerate(guide_data):
        row = guide_start_row + row_offset

        if row_offset == 0:  # Header row
            for col_offset, text in enumerate([color_name, meaning, threshold]):
                cell = worksheet.cell(row=row, column=guide_col + col_offset, value=text)
                cell.font = Font(bold=True, size=9)
                cell.fill = PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
        else:
            # Color cell
            cell = worksheet.cell(row=row, column=guide_col, value=color_name)
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal='center')
            if color_name == "Green":
                cell.fill = PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type="solid")
            elif color_name == "Orange":
                cell.fill = PatternFill(start_color=COLOR_ORANGE, end_color=COLOR_ORANGE, fill_type="solid")
            elif color_name == "Red":
                cell.fill = PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type="solid")

            # Meaning
            meaning_cell = worksheet.cell(row=row, column=guide_col + 1, value=meaning)
            meaning_cell.font = Font(size=9)

            # Threshold
            threshold_cell = worksheet.cell(row=row, column=guide_col + 2, value=threshold)
            threshold_cell.font = Font(size=9)

    # =========================================================================
    # SECTION 3: CHAPTER STATISTICS (LEFT SIDE)
    # =========================================================================

    # Section header
    worksheet.cell(row=current_row, column=1, value="Chapter Statistics").font = Font(bold=True, size=12)
    current_row += 1

    # Calculate performance tier percentages
    ref_tiers = count_performance_tiers(stats["ref_totals"], stats["avg_referrals"])
    oto_tiers = count_performance_tiers(stats["oto_totals"], stats["avg_oto"])
    tyfcb_tiers = count_performance_tiers(stats["tyfcb_totals"], stats["avg_tyfcb"])

    # Statistics table headers
    stat_headers = ["Metric", "Value"]
    for col_idx, header in enumerate(stat_headers, start=1):
        cell = worksheet.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type="solid")
    current_row += 1

    # Build statistics list
    statistics = [
        ("Chapter Size", stats["chapter_size"]),
        ("Period", period_str),
        ("Total Months", stats.get("total_months", 1)),
        ("", ""),
        ("Referrals - % Green", f"{ref_tiers['green_pct']:.1f}%"),
        ("Referrals - % Orange", f"{ref_tiers['orange_pct']:.1f}%"),
        ("Referrals - % Red", f"{ref_tiers['red_pct']:.1f}%"),
        ("", ""),
        ("OTO - % Green", f"{oto_tiers['green_pct']:.1f}%"),
        ("OTO - % Orange", f"{oto_tiers['orange_pct']:.1f}%"),
        ("OTO - % Red", f"{oto_tiers['red_pct']:.1f}%"),
        ("", ""),
        ("TYFCB - % Green", f"{tyfcb_tiers['green_pct']:.1f}%"),
        ("TYFCB - % Orange", f"{tyfcb_tiers['orange_pct']:.1f}%"),
        ("TYFCB - % Red", f"{tyfcb_tiers['red_pct']:.1f}%"),
    ]

    if differences:
        statistics.append(("", ""))
        statistics.append(("Inactive Members", len(differences)))

    # Write statistics with insights on the right
    stats_start_row = current_row
    for metric, value in statistics:
        worksheet.cell(row=current_row, column=1, value=metric).font = Font(bold=metric != "", size=10)
        worksheet.cell(row=current_row, column=2, value=value).font = Font(size=10)
        current_row += 1

    # =========================================================================
    # SECTION 4: CONTEXTUAL INSIGHTS (RIGHT SIDE - NEXT TO STATISTICS)
    # =========================================================================

    insights_col = 10
    insights_row = stats_start_row + 2  # Start a bit below statistics

    # Add contextual insights
    insights = []

    if ref_tiers['green_pct'] > 30:
        insights.append(f"✓ Strong referral performance: {ref_tiers['green_pct']:.0f}% excellent")
    elif ref_tiers['green_pct'] < 20:
        insights.append(f"⚠ Low referral performance: Only {ref_tiers['green_pct']:.0f}% excellent")

    if tyfcb_tiers['green_pct'] > 30:
        insights.append(f"✓ High TYFCB performance: {tyfcb_tiers['green_pct']:.0f}% excellent")

    if ref_tiers['red_pct'] > 30:
        insights.append(f"⚠ {ref_tiers['red_pct']:.0f}% members need coaching")

    if differences:
        insights.append(f"⚠ {len(differences)} members became inactive")

    # Write insights
    for insight in insights:
        worksheet.merge_cells(
            start_row=insights_row,
            start_column=insights_col,
            end_row=insights_row,
            end_column=insights_col + 2,
        )
        insight_cell = worksheet.cell(row=insights_row, column=insights_col)
        insight_cell.value = insight
        insight_cell.font = Font(size=9, italic=True)
        insight_cell.alignment = Alignment(wrap_text=True)
        insights_row += 1

    # =========================================================================
    # SECTION 5: TOP 5 PERFORMERS (LEFT SIDE)
    # =========================================================================

    current_row += 1

    # Section header
    worksheet.merge_cells(f'A{current_row}:E{current_row}')
    section_cell = worksheet.cell(row=current_row, column=1)
    section_cell.value = "⭐ TOP 5 PERFORMERS"
    section_cell.font = Font(bold=True, size=12)
    section_cell.fill = PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type='solid')
    section_cell.alignment = Alignment(horizontal='center')
    current_row += 1

    # Get all member performance data
    ref_matrix = aggregated_data["referral_matrix"]
    member_performance = []

    for member in ref_matrix.index:
        ref_val = stats["ref_totals"].get(member, 0)
        oto_val = stats["oto_totals"].get(member, 0)
        tyfcb_val = stats["tyfcb_totals"].get(member, 0)

        # Calculate normalized score
        ref_ratio = ref_val / stats["avg_referrals"] if stats["avg_referrals"] > 0 else 0
        oto_ratio = oto_val / stats["avg_oto"] if stats["avg_oto"] > 0 else 0
        tyfcb_ratio = tyfcb_val / stats["avg_tyfcb"] if stats["avg_tyfcb"] > 0 else 0
        total_score = ref_ratio + oto_ratio + tyfcb_ratio

        member_performance.append({
            "member": member,
            "score": total_score,
            "ref": ref_val,
            "oto": oto_val,
            "tyfcb": tyfcb_val,
        })

    # Sort by score descending
    member_performance.sort(key=lambda x: x["score"], reverse=True)

    # Top 5 headers
    top_headers = ["Rank", "Member", "Referrals", "OTO", "TYFCB (AED)"]
    for col_idx, header in enumerate(top_headers, start=1):
        cell = worksheet.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    current_row += 1

    # Write top 5
    top_5_start_row = current_row
    for rank, perf in enumerate(member_performance[:5], start=1):
        # Rank
        worksheet.cell(row=current_row, column=1, value=f"#{rank}").font = Font(bold=True, size=11)

        # Member name
        member_cell = worksheet.cell(row=current_row, column=2, value=perf["member"])
        member_cell.font = Font(bold=True, size=10)

        # Highlight #1 performer in green
        if rank == 1:
            for col in range(1, 6):
                cell = worksheet.cell(row=current_row, column=col)
                cell.fill = PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type='solid')

        # Referrals
        worksheet.cell(row=current_row, column=3, value=perf["ref"]).font = Font(size=10)

        # OTO
        worksheet.cell(row=current_row, column=4, value=perf["oto"]).font = Font(size=10)

        # TYFCB
        tyfcb_cell = worksheet.cell(row=current_row, column=5, value=perf["tyfcb"])
        tyfcb_cell.number_format = "#,##0.00"
        tyfcb_cell.font = Font(size=10)

        current_row += 1

    # =========================================================================
    # SECTION 6: TOP 5 INSIGHTS (RIGHT SIDE)
    # =========================================================================

    top_insights_col = 10
    top_insights_row = top_5_start_row

    # Add growth percentages for top performers
    for rank, perf in enumerate(member_performance[:5], start=1):
        # Calculate how much above average they are
        avg_score = 3.0  # Normalized average score is 3 (1+1+1)
        growth_pct = ((perf["score"] - avg_score) / avg_score * 100) if avg_score > 0 else 0

        if growth_pct > 0:
            insight_text = f"↑ {growth_pct:.0f}% above avg"
        else:
            insight_text = f"↓ {abs(growth_pct):.0f}% below avg"

        worksheet.merge_cells(
            start_row=top_insights_row,
            start_column=top_insights_col,
            end_row=top_insights_row,
            end_column=top_insights_col + 2,
        )
        insight_cell = worksheet.cell(row=top_insights_row, column=top_insights_col)
        insight_cell.value = insight_text
        insight_cell.font = Font(size=9, italic=True)
        insight_cell.alignment = Alignment(horizontal='center')
        top_insights_row += 1

    # =========================================================================
    # SECTION 7: BOTTOM 3 NEED ATTENTION (LEFT SIDE)
    # =========================================================================

    current_row += 1

    # Section header
    worksheet.merge_cells(f'A{current_row}:E{current_row}')
    section_cell = worksheet.cell(row=current_row, column=1)
    section_cell.value = "⚠️  BOTTOM 3 NEED ATTENTION"
    section_cell.font = Font(bold=True, size=12)
    section_cell.fill = PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type='solid')
    section_cell.alignment = Alignment(horizontal='center')
    current_row += 1

    # Bottom 3 headers
    bottom_headers = ["Member", "Referrals", "OTO", "TYFCB (AED)"]
    for col_idx, header in enumerate(bottom_headers, start=1):
        cell = worksheet.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    current_row += 1

    # Get bottom 3 (worst performers)
    bottom_3 = member_performance[-3:]
    bottom_3.reverse()  # Show worst first

    bottom_3_start_row = current_row
    for perf in bottom_3:
        # Member name
        member_cell = worksheet.cell(row=current_row, column=1, value=perf["member"])
        member_cell.font = Font(bold=True, size=10)

        # Referrals
        ref_cell = worksheet.cell(row=current_row, column=2, value=perf["ref"])
        ref_cell.font = Font(size=10)

        # OTO
        oto_cell = worksheet.cell(row=current_row, column=3, value=perf["oto"])
        oto_cell.font = Font(size=10)

        # TYFCB
        tyfcb_cell = worksheet.cell(row=current_row, column=4, value=perf["tyfcb"])
        tyfcb_cell.number_format = "#,##0.00"
        tyfcb_cell.font = Font(size=10)

        # Red background for all cells
        for col in range(1, 5):
            cell = worksheet.cell(row=current_row, column=col)
            cell.fill = PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type='solid')

        current_row += 1

    # =========================================================================
    # SECTION 8: BOTTOM 3 INSIGHTS (RIGHT SIDE)
    # =========================================================================

    bottom_insights_col = 10
    bottom_insights_row = bottom_3_start_row

    # Add drop percentages for bottom performers
    for perf in bottom_3:
        avg_score = 3.0  # Normalized average score
        drop_pct = ((avg_score - perf["score"]) / avg_score * 100) if avg_score > 0 else 0

        if drop_pct > 0:
            insight_text = f"↓ {drop_pct:.0f}% below avg"
        else:
            insight_text = f"↑ {abs(drop_pct):.0f}% above avg"

        worksheet.merge_cells(
            start_row=bottom_insights_row,
            start_column=bottom_insights_col,
            end_row=bottom_insights_row,
            end_column=bottom_insights_col + 2,
        )
        insight_cell = worksheet.cell(row=bottom_insights_row, column=bottom_insights_col)
        insight_cell.value = insight_text
        insight_cell.font = Font(size=9, italic=True)
        insight_cell.alignment = Alignment(horizontal='center')
        bottom_insights_row += 1

    # =========================================================================
    # SECTION 9: HYPERLINK BUTTON TO ALL MEMBERS TABLE
    # =========================================================================

    current_row += 2

    # Create hyperlink button
    worksheet.merge_cells(f'A{current_row}:E{current_row}')
    hyperlink_cell = worksheet.cell(row=current_row, column=1)
    hyperlink_cell.value = "📊 View All Members Report ↓"
    hyperlink_cell.font = Font(bold=True, size=12, color='0563C1', underline='single')
    hyperlink_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add hyperlink to row where full table starts (35 rows below current position)
    all_members_row = current_row + 35
    hyperlink_cell.hyperlink = f"#A{all_members_row}"

    # Make it look like a button
    hyperlink_cell.fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
    hyperlink_cell.border = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )
    worksheet.row_dimensions[current_row].height = 30

    # =========================================================================
    # SECTION 10: SPACER (35 empty rows)
    # =========================================================================

    current_row += 35

    # =========================================================================
    # SECTION 11: FULL ALL MEMBERS PERFORMANCE TABLE
    # =========================================================================

    # Section header
    worksheet.merge_cells(f'A{current_row}:F{current_row}')
    section_cell = worksheet.cell(row=current_row, column=1)
    section_cell.value = "ALL MEMBERS - COMPLETE PERFORMANCE DATA"
    section_cell.font = Font(bold=True, size=14)
    section_cell.fill = PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type='solid')
    section_cell.alignment = Alignment(horizontal='center')
    current_row += 1

    # Table headers
    perf_headers = ["Rank", "Member Name", "Referrals", "OTO", "TYFCB (AED)", "Overall Score"]
    for col_idx, header in enumerate(perf_headers, start=1):
        cell = worksheet.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    current_row += 1

    # Write all members with performance highlighting
    for rank, perf_data in enumerate(member_performance, start=1):
        # Rank
        worksheet.cell(row=current_row, column=1, value=rank).font = Font(bold=True, size=10)

        # Member name
        worksheet.cell(row=current_row, column=2, value=perf_data["member"]).font = Font(bold=True, size=10)

        # Referrals (with color)
        ref_cell = worksheet.cell(row=current_row, column=3, value=perf_data["ref"])
        ref_cell.font = Font(size=10)
        ref_color = get_performance_color(perf_data["ref"], stats["avg_referrals"])
        if ref_color:
            ref_cell.fill = PatternFill(start_color=ref_color, end_color=ref_color, fill_type="solid")
            ref_cell.font = Font(bold=True, size=10)

        # OTO (with color)
        oto_cell = worksheet.cell(row=current_row, column=4, value=perf_data["oto"])
        oto_cell.font = Font(size=10)
        oto_color = get_performance_color(perf_data["oto"], stats["avg_oto"])
        if oto_color:
            oto_cell.fill = PatternFill(start_color=oto_color, end_color=oto_color, fill_type="solid")
            oto_cell.font = Font(bold=True, size=10)

        # TYFCB (with color)
        tyfcb_cell = worksheet.cell(row=current_row, column=5, value=perf_data["tyfcb"])
        tyfcb_cell.number_format = "#,##0.00"
        tyfcb_cell.font = Font(size=10)
        tyfcb_color = get_performance_color(perf_data["tyfcb"], stats["avg_tyfcb"])
        if tyfcb_color:
            tyfcb_cell.fill = PatternFill(start_color=tyfcb_color, end_color=tyfcb_color, fill_type="solid")
            tyfcb_cell.font = Font(bold=True, size=10)

        # Overall Score
        score_cell = worksheet.cell(row=current_row, column=6, value=f"{perf_data['score']:.2f}")
        score_cell.font = Font(size=10)

        current_row += 1

    # =========================================================================
    # COLUMN WIDTHS
    # =========================================================================

    worksheet.column_dimensions['A'].width = 15
    worksheet.column_dimensions['B'].width = 25
    worksheet.column_dimensions['C'].width = 12
    worksheet.column_dimensions['D'].width = 12
    worksheet.column_dimensions['E'].width = 15
    worksheet.column_dimensions['F'].width = 15
    worksheet.column_dimensions['G'].width = 3  # Spacer
    worksheet.column_dimensions['H'].width = 3  # Spacer
    worksheet.column_dimensions['I'].width = 3  # Spacer

    # Insights/Guide columns (right side)
    worksheet.column_dimensions[get_column_letter(guide_col)].width = 12
    worksheet.column_dimensions[get_column_letter(guide_col + 1)].width = 15
    worksheet.column_dimensions[get_column_letter(guide_col + 2)].width = 20
