"""
Shared border utility functions for Excel reports.

Provides consistent border styling across all Excel sheets.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

from .colors import COLOR_HEADER_BG


def create_merged_header(worksheet, title: str, num_columns: int, period_str: str = None, row: int = 1):
    """
    Create merged header cell spanning multiple columns.

    Args:
        worksheet: The worksheet to add header to
        title: The title text (period will be appended if provided)
        num_columns: Number of columns to span
        period_str: Optional period string (e.g., "01/2025 - 03/2025")
        row: Row number for the header (default: 1)
    """
    # Build full title
    if period_str:
        full_title = f"{title} - Period: {period_str}"
    else:
        full_title = title

    # Merge cells
    end_col = get_column_letter(num_columns)
    worksheet.merge_cells(f"A{row}:{end_col}{row}")

    # Style the merged cell
    cell = worksheet[f"A{row}"]
    cell.value = full_title
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(
        start_color=COLOR_HEADER_BG,
        end_color=COLOR_HEADER_BG,
        fill_type="solid",
    )

    # Add border
    thick_border = Border(bottom=Side(style="thick"))
    cell.border = thick_border

    # Set fixed row height for merged header (prevents auto-resizing)
    worksheet.row_dimensions[row].height = 30


def apply_thin_borders(worksheet, start_row: int, end_row: int, start_col: int, end_col: int):
    """
    Apply thin borders to a range of cells for professional appearance.

    IMPORTANT: Call this LAST after adding thick borders to preserve them.

    Args:
        worksheet: The worksheet to add borders to
        start_row: Starting row number
        end_row: Ending row number
        start_col: Starting column number
        end_col: Ending column number
    """
    thin_side = Side(style="thin", color="000000")

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            existing = cell.border

            # Determine which border sides to use
            left_side = thin_side
            right_side = thin_side
            top_side = thin_side
            bottom_side = thin_side

            # Keep medium/thick borders if they exist
            if (
                existing
                and existing.left
                and hasattr(existing.left, "style")
                and existing.left.style in ["medium", "thick"]
            ):
                left_side = existing.left
            if (
                existing
                and existing.right
                and hasattr(existing.right, "style")
                and existing.right.style in ["medium", "thick"]
            ):
                right_side = existing.right
            if (
                existing
                and existing.top
                and hasattr(existing.top, "style")
                and existing.top.style in ["medium", "thick"]
            ):
                top_side = existing.top
            if (
                existing
                and existing.bottom
                and hasattr(existing.bottom, "style")
                and existing.bottom.style in ["medium", "thick"]
            ):
                bottom_side = existing.bottom

            cell.border = Border(
                left=left_side,
                right=right_side,
                top=top_side,
                bottom=bottom_side,
            )


def add_thick_right_border(worksheet, col_idx: int, start_row: int, end_row: int):
    """
    Add thick right border to a column for section separation.

    Args:
        worksheet: The worksheet to add border to
        col_idx: Column index (1-based)
        start_row: Starting row number
        end_row: Ending row number
    """
    for row in range(start_row, end_row + 1):
        cell = worksheet.cell(row=row, column=col_idx)
        # Combine with existing borders
        existing = cell.border
        cell.border = Border(
            left=existing.left if existing else None,
            right=Side(style="medium", color="000000"),
            top=existing.top if existing else None,
            bottom=existing.bottom if existing else None,
        )


def add_bottom_border_to_row(worksheet, row_idx: int, start_col: int, end_col: int, style: str = "medium"):
    """
    Add bottom border to a row (useful for headers).

    Args:
        worksheet: The worksheet to add border to
        row_idx: Row index (1-based)
        start_col: Starting column number
        end_col: Ending column number
        style: Border style ('thin', 'medium', 'thick')
    """
    for col in range(start_col, end_col + 1):
        cell = worksheet.cell(row=row_idx, column=col)
        existing = cell.border
        cell.border = Border(
            left=existing.left if existing else None,
            right=existing.right if existing else None,
            top=existing.top if existing else None,
            bottom=Side(style=style, color="000000"),
        )


def add_outer_table_borders(worksheet, start_row: int, end_row: int, start_col: int, end_col: int):
    """
    Add medium borders around the outer edges of a table for a complete boxed look.

    Args:
        worksheet: The worksheet to add borders to
        start_row: Starting row number
        end_row: Ending row number
        start_col: Starting column number
        end_col: Ending column number
    """
    medium_side = Side(style="medium", color="000000")

    # Top edge
    for col in range(start_col, end_col + 1):
        cell = worksheet.cell(row=start_row, column=col)
        existing = cell.border
        cell.border = Border(
            top=medium_side,
            left=existing.left if existing else None,
            right=existing.right if existing else None,
            bottom=existing.bottom if existing else None,
        )

    # Bottom edge
    for col in range(start_col, end_col + 1):
        cell = worksheet.cell(row=end_row, column=col)
        existing = cell.border
        cell.border = Border(
            bottom=medium_side,
            left=existing.left if existing else None,
            right=existing.right if existing else None,
            top=existing.top if existing else None,
        )

    # Left edge
    for row in range(start_row, end_row + 1):
        cell = worksheet.cell(row=row, column=start_col)
        existing = cell.border
        cell.border = Border(
            left=medium_side,
            top=existing.top if existing else None,
            right=existing.right if existing else None,
            bottom=existing.bottom if existing else None,
        )

    # Right edge
    for row in range(start_row, end_row + 1):
        cell = worksheet.cell(row=row, column=end_col)
        existing = cell.border
        cell.border = Border(
            right=medium_side,
            top=existing.top if existing else None,
            left=existing.left if existing else None,
            bottom=existing.bottom if existing else None,
        )


def apply_standard_table_borders(worksheet, start_row: int, end_row: int, start_col: int, end_col: int,
                                 thick_separator_cols: list = None):
    """
    Apply standard professional borders to a table.

    This is a convenience function that applies borders in the correct order:
    1. Outer table borders (medium)
    2. Bottom border for header row (medium)
    3. Thick right borders for section separators
    4. Thin borders for all cells (preserving thick borders)

    Args:
        worksheet: The worksheet to add borders to
        start_row: Starting row number (usually row 2 for headers)
        end_row: Ending row number (including total row)
        start_col: Starting column number
        end_col: Ending column number
        thick_separator_cols: List of column indices that should have thick right borders
    """
    # 1. Outer table borders first
    add_outer_table_borders(worksheet, start_row, end_row, start_col, end_col)

    # 2. Medium border below headers
    add_bottom_border_to_row(worksheet, start_row, start_col, end_col, style="medium")

    # 3. Thick right borders for section separators
    if thick_separator_cols:
        for col_idx in thick_separator_cols:
            add_thick_right_border(worksheet, col_idx, start_row, end_row)

    # 4. Thin borders to ALL cells LAST (fills in the grid)
    apply_thin_borders(worksheet, start_row, end_row, start_col, end_col)
