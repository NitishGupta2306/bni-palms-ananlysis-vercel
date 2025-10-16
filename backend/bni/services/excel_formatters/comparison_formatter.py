"""
Comparison/Inactive Members Excel Formatter.

Handles formatting for member differences report showing:
- Members who became inactive during the period
- Last active month for each inactive member
- Business name and classification
- Professional table formatting with borders
"""

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .colors import COLOR_GRAY
from .border_utils import apply_standard_table_borders


def write_inactive_members(worksheet, differences: list, period_str: str):
    """
    Write member differences (inactive members) to worksheet.

    Args:
        worksheet: The worksheet to write to
        differences: List of dicts with member info:
            - member_name: Full name
            - business_name: Business name
            - classification: Member classification
            - last_active_month: Last month they were active
        period_str: Period display string (e.g., "01/2025 - 03/2025")
    """
    # Title
    worksheet["A1"] = "Inactive Members Report"
    worksheet["A1"].font = Font(bold=True, size=14)

    worksheet["A2"] = f"Period: {period_str}"

    # Headers
    headers = [
        "Member Name",
        "Business Name",
        "Classification",
        "Last Active Month",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type="solid"
        )

    # Data
    last_row = 5
    for row_idx, member_data in enumerate(differences, start=5):
        worksheet.cell(row=row_idx, column=1, value=member_data["member_name"])
        worksheet.cell(
            row=row_idx, column=2, value=member_data.get("business_name", "")
        )
        worksheet.cell(
            row=row_idx, column=3, value=member_data.get("classification", "")
        )
        worksheet.cell(
            row=row_idx, column=4, value=member_data["last_active_month"]
        )
        last_row = row_idx

    # Add professional borders if we have data
    if differences:
        # Apply standard borders (outer, header bottom, no thick separators for this simple table, thin grid)
        apply_standard_table_borders(
            worksheet, 4, last_row, 1, 4, thick_separator_cols=None
        )

    # Set default column widths
    worksheet.column_dimensions["A"].width = 25  # Member name
    worksheet.column_dimensions["B"].width = 25  # Business name
    worksheet.column_dimensions["C"].width = 20  # Classification
    worksheet.column_dimensions["D"].width = 18  # Last active month
