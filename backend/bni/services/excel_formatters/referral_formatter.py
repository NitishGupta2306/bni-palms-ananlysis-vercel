"""
Referral Matrix Excel Formatter.

Handles formatting for referral matrix sheets with:
- Merged header with period display
- Rotated column headers for member names
- Performance highlighting (Green/Orange/Red) on member names and totals
- Yellow highlighting on non-zero referral values
- Monthly breakdown columns (Total + Unique per month)
- Professional borders with section separators
"""

import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .colors import (
    COLOR_YELLOW,
    COLOR_GRAY,
    get_performance_color,
)
from .border_utils import (
    create_merged_header,
    apply_standard_table_borders,
    add_thick_right_border,
)


def write_referral_matrix(worksheet, aggregated_matrix, period_str: str, stats: dict, reports: list):
    """
    Write referral matrix with monthly breakdowns and performance highlighting.

    Args:
        worksheet: The worksheet to write to
        aggregated_matrix: DataFrame or dict with aggregated referral data
        period_str: Period display string (e.g., "01/2025 - 03/2025")
        stats: Chapter statistics dict with keys:
            - ref_totals: Dict of {member_name: total_referrals}
            - avg_referrals: Chapter average referrals
        reports: List of MonthlyReport objects for monthly breakdowns
    """
    # Convert to DataFrame if needed
    if isinstance(aggregated_matrix, dict):
        df = pd.DataFrame(aggregated_matrix)
    else:
        df = aggregated_matrix

    num_members = len(df.columns)
    num_months = len(reports)

    # For single-month reports, skip monthly breakdowns (would just duplicate aggregates)
    show_monthly_breakdown = num_months > 1

    # Calculate total columns: 1 (Member col) + num_members + 2 aggregates + (2 monthly cols) * num_months
    if show_monthly_breakdown:
        total_columns = 1 + num_members + 2 + (num_months * 2)
    else:
        total_columns = 1 + num_members + 2  # No monthly columns for single month

    # Create merged header
    create_merged_header(worksheet, "Referral Matrix", total_columns, period_str, row=1)

    # Freeze panes at row 3 (below headers)
    worksheet.freeze_panes = "B3"

    # =========================================================================
    # COLUMN HEADERS (Row 2)
    # =========================================================================

    # Member column header
    cell = worksheet.cell(row=2, column=1, value="Member")
    cell.font = Font(bold=True)
    cell.fill = PatternFill(
        start_color=COLOR_GRAY,
        end_color=COLOR_GRAY,
        fill_type="solid",
    )
    cell.alignment = Alignment(horizontal="center", vertical="center")

    current_col = 2

    # Member name column headers (rotated 90°)
    for member_name in df.columns:
        cell = worksheet.cell(row=2, column=current_col)
        cell.value = member_name
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(
            textRotation=90, horizontal="center", vertical="bottom"
        )
        cell.fill = PatternFill(
            start_color=COLOR_GRAY,
            end_color=COLOR_GRAY,
            fill_type="solid",
        )
        current_col += 1

    # Aggregate column headers
    agg_start_col = current_col
    worksheet.cell(row=2, column=current_col, value="Total Given").font = Font(bold=True)
    worksheet.cell(row=2, column=current_col, value="Total Given").fill = PatternFill(
        start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type="solid"
    )
    current_col += 1

    worksheet.cell(row=2, column=current_col, value="Unique Given").font = Font(bold=True)
    worksheet.cell(row=2, column=current_col, value="Unique Given").fill = PatternFill(
        start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type="solid"
    )
    agg_end_col = current_col
    current_col += 1

    # Monthly column headers (only for multi-month reports)
    if show_monthly_breakdown:
        for idx, report in enumerate(reports, start=1):
            month_date = datetime.strptime(report.month_year, "%Y-%m")
            month_display = month_date.strftime("%m/%Y")

            cell = worksheet.cell(
                row=2, column=current_col, value=f"M{idx}-{month_display}\nTotal"
            )
            cell.font = Font(bold=True, size=8)
            cell.alignment = Alignment(
                textRotation=90, horizontal="center", vertical="bottom", wrap_text=True
            )
            cell.fill = PatternFill(
                start_color=COLOR_GRAY,
                end_color=COLOR_GRAY,
                fill_type="solid",
            )
            current_col += 1

            cell = worksheet.cell(
                row=2, column=current_col, value=f"M{idx}-{month_display}\nUnique"
            )
            cell.font = Font(bold=True, size=8)
            cell.alignment = Alignment(
                textRotation=90, horizontal="center", vertical="bottom", wrap_text=True
            )
            cell.fill = PatternFill(
                start_color=COLOR_GRAY,
                end_color=COLOR_GRAY,
                fill_type="solid",
            )
            current_col += 1

    # Set header row height for rotated text visibility
    worksheet.row_dimensions[2].height = 60

    # =========================================================================
    # DATA ROWS (Starting at row 3)
    # =========================================================================

    member_totals = stats["ref_totals"]
    avg_value = stats["avg_referrals"]
    yellow_fill = PatternFill(
        start_color=COLOR_YELLOW,
        end_color=COLOR_YELLOW,
        fill_type="solid",
    )

    current_row = 3
    for row_name in df.index:
        # Member name (with performance highlighting)
        member_cell = worksheet.cell(row=current_row, column=1, value=row_name)
        member_cell.font = Font(bold=True)

        # Apply performance color to member name
        perf_color = get_performance_color(
            member_totals.get(row_name, 0), avg_value
        )
        if perf_color:
            member_cell.fill = PatternFill(
                start_color=perf_color, end_color=perf_color, fill_type="solid"
            )

        # Matrix data cells
        col_idx = 2
        for col_name in df.columns:
            value = df.loc[row_name, col_name]
            cell = worksheet.cell(row=current_row, column=col_idx, value=value)

            # Yellow highlight for non-zero values
            if value and value > 0:
                cell.fill = yellow_fill
                cell.font = Font(bold=True)

            col_idx += 1

        # Calculate aggregate totals for this member
        row_total = df.loc[row_name].sum()
        unique_count = (df.loc[row_name] > 0).sum()

        # Total Given (with performance highlighting)
        total_cell = worksheet.cell(
            row=current_row, column=agg_start_col, value=row_total
        )
        total_cell.font = Font(bold=True)
        if perf_color:
            total_cell.fill = PatternFill(
                start_color=perf_color, end_color=perf_color, fill_type="solid"
            )

        # Unique Given (with performance highlighting)
        unique_cell = worksheet.cell(
            row=current_row, column=agg_start_col + 1, value=unique_count
        )
        unique_cell.font = Font(bold=True)
        if perf_color:
            unique_cell.fill = PatternFill(
                start_color=perf_color, end_color=perf_color, fill_type="solid"
            )

        # Monthly totals (only for multi-month reports)
        if show_monthly_breakdown:
            # Move to monthly columns (right after aggregate columns)
            col_idx = agg_end_col + 1

            for report in reports:
                month_matrix_data = report.referral_matrix_data

                # Calculate monthly totals for this member
                if (
                    month_matrix_data
                    and "members" in month_matrix_data
                    and "matrix" in month_matrix_data
                ):
                    members = month_matrix_data["members"]
                    matrix = month_matrix_data["matrix"]

                    if row_name in members:
                        member_idx = members.index(row_name)
                        if member_idx < len(matrix):
                            month_row = matrix[member_idx]
                            month_total = sum(
                                val
                                for val in month_row
                                if isinstance(val, (int, float)) and val > 0
                            )
                            month_unique = sum(
                                1
                                for val in month_row
                                if isinstance(val, (int, float)) and val > 0
                            )

                            worksheet.cell(
                                row=current_row, column=col_idx, value=month_total
                            )
                            worksheet.cell(
                                row=current_row, column=col_idx + 1, value=month_unique
                            )
                        else:
                            worksheet.cell(row=current_row, column=col_idx, value=0)
                            worksheet.cell(row=current_row, column=col_idx + 1, value=0)
                    else:
                        worksheet.cell(row=current_row, column=col_idx, value=0)
                        worksheet.cell(row=current_row, column=col_idx + 1, value=0)
                else:
                    worksheet.cell(row=current_row, column=col_idx, value=0)
                    worksheet.cell(row=current_row, column=col_idx + 1, value=0)

                col_idx += 2  # Move to next month

        current_row += 1

    # =========================================================================
    # TOTAL RECEIVED ROW
    # =========================================================================

    total_row = current_row
    worksheet.cell(row=total_row, column=1, value="Total Received").font = Font(
        bold=True
    )

    col_idx = 2
    for col_name in df.columns:
        col_total = df[col_name].sum()
        cell = worksheet.cell(row=total_row, column=col_idx, value=col_total)
        cell.font = Font(bold=True)
        col_idx += 1

    # =========================================================================
    # BORDERS (Order matters!)
    # =========================================================================

    # Calculate thick separator positions
    thick_separator_cols = [agg_end_col]  # After "Unique Given" column

    # Add monthly separators only for multi-month reports
    if show_monthly_breakdown:
        month_col = agg_end_col + 1
        for idx in range(num_months - 1):
            month_end = month_col + 1  # After each month's "Unique" column
            thick_separator_cols.append(month_end)
            month_col += 2

    # Apply standard borders (outer, header bottom, thick separators, thin grid)
    apply_standard_table_borders(
        worksheet, 2, total_row, 1, total_columns, thick_separator_cols
    )

    # =========================================================================
    # COLUMN WIDTHS
    # =========================================================================

    # Member name column (first column)
    worksheet.column_dimensions["A"].width = 20

    # Member name columns (matrix columns - narrower for rotated text)
    for col in range(2, 2 + num_members):
        worksheet.column_dimensions[get_column_letter(col)].width = 4

    # Aggregate columns (Total Given, Unique Given)
    for col in range(agg_start_col, agg_end_col + 1):
        worksheet.column_dimensions[get_column_letter(col)].width = 15

    # Monthly columns
    if show_monthly_breakdown:
        for col in range(agg_end_col + 1, total_columns + 1):
            worksheet.column_dimensions[get_column_letter(col)].width = 10
